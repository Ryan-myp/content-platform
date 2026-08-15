#!/usr/bin/env python3
"""图片模板商城 — 图片创意平台的商业化核心。

能力：
1. 模板定价：模板 JSON 内嵌 pricing 字段
   {"mode": "free"|"once"|"day"|"month", "once": N, "day": N, "month": N}（积分）
   - once：按次买断，永久使用；day：按天订阅；month：按月订阅
2. 购买/订阅：从 user_quotas.credits 扣减，写 image_template_access
3. 渲染鉴权：收费模板渲染前校验权限（image_factory 渲染 worker 接入 check_render_access）
4. 热度统计：渲染成功记 image_template_stats.usage_count（市场页「使用次数」）
5. Excel 批量生成：上传 xlsx/csv + 字段映射 → 逐行套版渲染 → zip 打包下载
   （对标 Shopee/跨境卖家「Excel 批量出图」工具场景）

表结构：
- image_template_access  (购买/订阅记录：once 永久；day/month 带过期时间)
- image_template_stats   (使用热度)
"""

import csv
import json
import logging
import os
import uuid
import zipfile
from datetime import datetime, timedelta

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel

from common.auth import require_auth
from common.helpers import _notify_progress
from common.config import load_config
from common.db import get_db
from common.llm import _safe_exc_msg
from task_queue import create_task, register_handler

logger = logging.getLogger(__name__)
load_config()

router = APIRouter(prefix="/api/image-store", tags=["图片模板商城"])

BATCH_DIR = os.path.join(os.path.dirname(__file__), "uploads", "batch")
os.makedirs(BATCH_DIR, exist_ok=True)

FREE_PRICING = {"mode": "free", "once": 0, "day": 0, "month": 0}
MODE_LABELS = {"free": "免费", "once": "按次", "day": "按天", "month": "按月"}


# ══════════════════════════════════════════════════════════════
# 数据层
# ══════════════════════════════════════════════════════════════


def _ensure_tables(conn) -> None:
    conn.execute(
        """CREATE TABLE IF NOT EXISTS image_template_access (
            id TEXT PRIMARY KEY,
            user_id TEXT DEFAULT '',
            template_id TEXT NOT NULL,
            access_type TEXT DEFAULT 'once',
            price INTEGER DEFAULT 0,
            expires_at TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS image_template_stats (
            template_id TEXT PRIMARY KEY,
            usage_count INTEGER DEFAULT 0,
            last_used_at TEXT DEFAULT ''
        )"""
    )
    conn.commit()


def get_pricing(template: dict) -> dict:
    """规范化模板 pricing 字段（缺省为免费）。"""
    p = template.get("pricing") or {}
    if not isinstance(p, dict):
        p = {}
    mode = p.get("mode", "free")
    if mode not in ("free", "once", "day", "month"):
        mode = "free"
    return {
        "mode": mode,
        "once": max(0, int(p.get("once", 0) or 0)),
        "day": max(0, int(p.get("day", 0) or 0)),
        "month": max(0, int(p.get("month", 0) or 0)),
    }


def is_paid(pricing: dict) -> bool:
    return pricing["mode"] != "free"


def _template_dir() -> str:
    # 延迟导入避免循环依赖（image_factory 顶部会 import 本模块的鉴权函数）
    from image_factory import TEMPLATE_DIR

    return TEMPLATE_DIR


def load_template(template_id: str) -> dict:
    path = os.path.join(_template_dir(), f"{template_id}.json")
    if not os.path.exists(path):
        raise HTTPException(404, "模板不存在")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _get_usage(template_id: str) -> int:
    try:
        conn = get_db()
        _ensure_tables(conn)
        row = conn.execute(
            "SELECT usage_count FROM image_template_stats WHERE template_id=?", (template_id,)
        ).fetchone()
        conn.close()
        return int(row["usage_count"]) if row else 0
    except Exception:
        return 0


def get_usage_stats() -> dict:
    """返回所有模板的使用热度统计：{template_id: usage_count}。"""
    result = {}
    try:
        conn = get_db()
        _ensure_tables(conn)
        rows = conn.execute(
            "SELECT template_id, usage_count FROM image_template_stats ORDER BY usage_count DESC"
        ).fetchall()
        conn.close()
        for r in rows:
            result[r["template_id"]] = int(r["usage_count"])
    except Exception:
        pass
    return result


def user_access(conn, username: str, template_id: str) -> dict | None:
    """用户对模板的有效权限（once 永久；day/month 未过期返回最新一条）。"""
    rows = conn.execute(
        """SELECT * FROM image_template_access
           WHERE user_id=? AND template_id=? ORDER BY created_at DESC""",
        (username, template_id),
    ).fetchall()
    now = datetime.now()
    for r in rows:
        a = dict(r)
        if a["access_type"] == "once":
            return a
        if a["expires_at"]:
            try:
                if datetime.fromisoformat(a["expires_at"]) > now:
                    return a
            except ValueError:
                pass
    return None


def check_render_access(username: str, template: dict) -> dict:
    """渲染鉴权（本地免费版：无计费，全部模板直接可用）。"""
    return get_pricing(template)  # 本地版不收费，所有模板可直接渲染


def record_usage(template_id: str) -> None:
    """模板使用热度 +1（渲染成功后调用，批量按行数累计）。"""
    try:
        conn = get_db()
        _ensure_tables(conn)
        now = datetime.now().isoformat()
        conn.execute(
            """INSERT INTO image_template_stats (template_id, usage_count, last_used_at)
               VALUES (?,1,?)
               ON CONFLICT(template_id) DO UPDATE SET usage_count=usage_count+1, last_used_at=? """,
            (template_id, now, now),
        )
        conn.commit()
        conn.close()
    except Exception as e:  # 统计失败不影响主流程
        logger.debug(f"record_usage skipped: {e}")


def _deduct_credits(conn, username: str, amount: int) -> None:
    """扣积分（余额不足抛 402）。"""
    quota = conn.execute(
        "SELECT credits FROM user_quotas WHERE username=?", (username,)
    ).fetchone()
    balance = int(quota["credits"]) if quota else 0
    if balance < amount:
        raise HTTPException(402, "余额不足，请先充值")
    conn.execute(
        "UPDATE user_quotas SET credits=credits-?, updated_at=? WHERE username=?",
        (amount, datetime.now().isoformat(), username),
    )


# ══════════════════════════════════════════════════════════════
# 市场 API
# ══════════════════════════════════════════════════════════════


class PurchaseRequest(BaseModel):
    access_type: str = "once"  # once | day | month



def _market_load_templates() -> list:
    """加载市场模板列表（跳过隐藏项）。"""
    tdir = _template_dir()
    items = []
    if os.path.exists(tdir):
        for f in os.listdir(tdir):
            if not f.endswith(".json"):
                continue
            try:
                with open(os.path.join(tdir, f), encoding="utf-8") as fh:
                    t = json.load(fh)
            except Exception:
                continue
            tid = t.get("id", "")
            if not tid or t.get("hidden"):
                continue
            pricing = get_pricing(t)
            items.append({
                "id": tid,
                "name": t.get("name", "未命名模板"),
                "category": t.get("category", "通用") or "通用",
                "width": t.get("width", 1080),
                "height": t.get("height", 1920),
                "preview": f"/api/image-factory/template-preview/{tid}",
                "pricing": pricing,
                "pricing_label": MODE_LABELS[pricing["mode"]],
                "seller": t.get("seller", "platform"),
                "usage": _get_usage(tid),
                "created_at": t.get("created_at", ""),
            })
    return items


def _market_access_map(user: str) -> dict:
    """查询用户对模板的访问权限映射。"""
    conn = get_db()
    _ensure_tables(conn)
    access_rows = conn.execute(
        "SELECT * FROM image_template_access WHERE user_id=?", (user,)
    ).fetchall()
    conn.close()
    owned: dict = {}
    for r in access_rows:
        a = dict(r)
        if a["access_type"] == "once":
            owned[a["template_id"]] = a["access_type"]
        elif a["expires_at"]:
            try:
                if datetime.fromisoformat(a["expires_at"]) > datetime.now():
                    owned.setdefault(a["template_id"], a["access_type"])
            except ValueError:
                pass
    return owned


def _market_filter_sort(items: list, q: str, category: str, sort: str) -> list:
    """市场列表搜索/分类过滤 + 排序。"""
    if q:
        ql = q.strip().lower()
        items = [i for i in items if ql in i["name"].lower() or ql in i["category"].lower()]
    if category and category != "全部":
        items = [i for i in items if i["category"] == category]
    if sort == "new":
        items.sort(key=lambda i: i.get("created_at", ""), reverse=True)
    elif sort == "price":
        items.sort(key=lambda i: (0 if i["pricing"]["mode"] == "free" else 1, i["pricing"]["once"]))
    else:
        items.sort(key=lambda i: (i["usage"], i["created_at"]), reverse=True)
    return items


def _market_categories(items: list) -> list:
    """分类聚合（市场分类 tab）。"""
    cats: dict = {}
    for it in items:
        c = it["category"]
        cats.setdefault(c, {"label": c, "count": 0})
        cats[c]["count"] += 1
    return list(cats.values())

@router.get("/list")
async def market_list(
    category: str = "",
    q: str = "",
    sort: str = "hot",  # hot=热度 | new=最新 | price=价格
    current_user: dict = require_auth(),
):
    """图片模板市场列表（内置 + 用户创作模板聚合）。

    每项含：预览图、分类、定价（mode/价格）、热度、销量、我的权限状态。
    """
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    items = _market_load_templates()

    # 我的权限状态
    owned = _market_access_map(user)
    for it in items:
        it["access"] = owned.get(it["id"], "")

    # 搜索/分类过滤 + 排序
    items = _market_filter_sort(items, q, category, sort)

    # 分类聚合 + 用户积分余额
    cats = _market_categories(items)
    conn = get_db()
    quota = conn.execute("SELECT credits FROM user_quotas WHERE username=?", (user,)).fetchone()
    conn.close()
    return {
        "total": len(items),
        "items": items,
        "categories": cats,
        "credits": int(quota["credits"]) if quota else 0,
    }


@router.get("/my-access")
async def my_access(current_user: dict = require_auth()):
    """我的模板购买/订阅记录（含到期时间）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_tables(conn)
    rows = conn.execute(
        """SELECT * FROM image_template_access WHERE user_id=?
           ORDER BY created_at DESC LIMIT 200""",
        (user,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.post("/templates/{template_id}/purchase")
async def purchase_template(
    template_id: str, req: PurchaseRequest, current_user: dict = require_auth()
):
    """购买/订阅模板：once 永久；day 1 天；month 30 天。从积分余额扣减。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    if req.access_type not in ("once", "day", "month"):
        raise HTTPException(400, "access_type 仅支持 once/day/month")

    template = load_template(template_id)
    pricing = get_pricing(template)
    if not is_paid(pricing):
        return {"owned": True, "price_paid": 0, "message": "该模板免费，可直接使用"}

    price = pricing[req.access_type]
    if price <= 0:
        raise HTTPException(400, "操作失败，请稍后重试")

    conn = get_db()
    _ensure_tables(conn)
    # 幂等：once 已买永久有效；day/month 未过期则顺延到期
    existing = user_access(conn, user, template_id)
    if existing and (existing["access_type"] == "once" or req.access_type == "once"):
        conn.close()
        return {"owned": True, "price_paid": 0, "message": "你已拥有该模板权限，无需重复购买"}

    try:
        _deduct_credits(conn, user, price)
        now = datetime.now()
        if req.access_type == "once":
            expires = ""
        elif req.access_type == "day":
            base = (
                datetime.fromisoformat(existing["expires_at"])
                if existing and existing["expires_at"]
                else now
            )
            expires = (base + timedelta(days=1)).isoformat()
        else:
            base = (
                datetime.fromisoformat(existing["expires_at"])
                if existing and existing["expires_at"]
                else now
            )
            expires = (base + timedelta(days=30)).isoformat()
        aid = f"imga_{uuid.uuid4().hex[:10]}"
        conn.execute(
            """INSERT INTO image_template_access
               (id, user_id, template_id, access_type, price, expires_at, created_at)
               VALUES (?,?,?,?,?,?,?)""",
            (aid, user, template_id, req.access_type, price, expires, now.isoformat()),
        )
        conn.commit()
    finally:
        conn.close()
    return {
        "owned": True,
        "access_type": req.access_type,
        "price_paid": price,
        "expires_at": expires,
        "message": f"购买成功！{MODE_LABELS[req.access_type]}模式，花费 {price} 积分"
        + (f"，{expires[:10]} 到期" if expires else "，永久有效"),
    }


# ══════════════════════════════════════════════════════════════
# Excel 批量生成（对标 Shopee 卖家批量套版工具）
# ══════════════════════════════════════════════════════════════


def _parse_header(filepath: str, filename: str) -> list[str]:
    """只读取表头列名（xlsx/csv），供批量字段映射 UI 使用，避免解析全表。"""
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        with open(filepath, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
            except StopIteration:
                return []
        return [str(h).strip() for h in header]
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(400, "服务端未安装 openpyxl，无法解析 Excel") from e
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    try:
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        row = None
    wb.close()
    if not row:
        return []
    return [str(c).strip() if c is not None else "" for c in row]


def _parse_table(filepath: str, filename: str) -> list[dict]:
    """解析 xlsx/csv → 行 dict 列表（首行为表头）。"""
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        rows = []
        with open(filepath, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append({k.strip(): (v or "").strip() for k, v in r.items()})
        return rows
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(400, "服务端未安装 openpyxl，无法解析 Excel") from e
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(
            {
                header[i]: (str(v).strip() if v is not None else "")
                for i, v in enumerate(row)
                if i < len(header) and header[i]
            }
        )
    wb.close()
    return rows


def _layer_types(template: dict) -> dict:
    """key → 图层类型 映射（用于判断字段填的是文字还是图片）。"""
    m = {}
    for layer in template.get("layers", []):
        k = layer.get("key", "")
        if k:
            m[k] = layer.get("type", "text")
    return m



def _build_row_overrides(row: dict, field_map: dict, layer_types: dict) -> dict:
    """按字段映射构建单行渲染参数（图片/文本分离）。"""
    overrides = {}
    images: dict = {}
    for col, key in field_map.items():
        val = row.get(col, "")
        if not key:
            continue
        if layer_types.get(key) == "image":
            if val:
                images[key] = val
        else:
            overrides[key] = val
    if images:
        overrides["images"] = images
    return overrides


async def _render_batch_rows(rows: list, field_map: dict, layer_types: dict, template, total: int, _report, template_id: str) -> list:
    """逐行渲染模板，返回图片 URL 列表。"""
    from image_factory import save_image, render_template_image

    results = []
    for i, row in enumerate(rows):
        _report(5 + int(i * 85 / total), f"正在生成第 {i + 1}/{total} 张（{row.get(next(iter(row), ''), '')}）")
        overrides = _build_row_overrides(row, field_map, layer_types)
        try:
            imgs = await render_template_image(template, overrides)
        except Exception as e:
            logger.warning(f"批量渲染第 {i + 1} 行失败: {e}")
            raise HTTPException(500, "操作失败，请稍后重试")
        fname = save_image(imgs[0])
        results.append(f"/api/image-factory/images/{fname}")
        record_usage(template_id)
    return results


def _zip_batch_results(results: list, rows: list, batch_name: str, task_id: str) -> str:
    """批量结果 zip 打包 + 生成清单。"""
    import zipfile

    zip_name = f"batch_{task_id}_{int(datetime.now().timestamp())}.zip"
    zip_path = os.path.join(BATCH_DIR, zip_name)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, url in enumerate(results):
            fname = url.rsplit("/", 1)[-1]
            src = os.path.join(IMAGE_DIR, fname)
            if os.path.exists(src):
                zf.write(src, f"{i + 1:03d}_{batch_name}_{fname}")
    manifest = "\n".join(
        f"{i + 1}\t{row.get(next(iter(row), ''), '')}\t{url}" for i, (row, url) in enumerate(zip(rows, results))
    )
    with zipfile.ZipFile(zip_path, "a") as zf:
        zf.writestr("生成清单.tsv", f"序号\t首列值\t图片地址\n{manifest}")
    return zip_path

async def _image_batch_worker(task_id: str, payload: dict, update, ctx: dict) -> dict:
    """Excel/CSV 批量套版：逐行渲染 → 保存 → zip 打包。"""
    from image_factory import IMAGE_DIR, render_template_image, save_image

    username = ctx.get("username", "") or payload.get("username", "")
    template_id = payload.get("template_id", "")
    filepath = payload.get("filepath", "")
    field_map = payload.get("field_map") or {}
    batch_name = (payload.get("batch_name") or "批量生成").strip()[:40]

    template = load_template(template_id)
    # 鉴权：收费模板需有效权限（once 永久；day/month 未过期）；按次模板额外按行扣积分
    pricing = check_render_access(username, template)
    rows = _parse_table(filepath, payload.get("filename", "data.xlsx"))
    if not rows:
        raise HTTPException(400, "表格为空或没有数据行（请保留表头）")
    # 校验字段映射有效性
    layer_types = _layer_types(template)
    unknown = [k for k in field_map.values() if k and k not in layer_types]
    if unknown:
        raise HTTPException(400, "操作失败，请稍后重试")
    total = len(rows)
    if total > 500:
        raise HTTPException(400, "单次导入最多500行，请分批导入")

    # 按次模板：批量视为 N 次使用，逐行扣费（余额不足中止）
    if pricing["mode"] == "once" and pricing["once"] > 0:
        conn = get_db()
        _ensure_tables(conn)
        try:
            _deduct_credits(conn, username, pricing["once"] * total)
            conn.commit()
        finally:
            conn.close()

    def _report(pct, stage):
        _notify_progress(update, pct, stage)

    results = await _render_batch_rows(rows, field_map, layer_types, template, total, _report, template_id)

    # zip 打包
    zip_path = _zip_batch_results(results, rows, batch_name, task_id)

    _report(100, f"批量生成完成，共 {total} 张")
    return {
        "count": total,
        "images": results,
        "zip": f"/api/image-store/batch/{task_id}/download",
        "zip_path": zip_path,
    }


async def _image_batch_handler(task_id: str, payload: dict, update, ctx: dict) -> dict:
    return await _image_batch_worker(task_id, payload, update, ctx)


register_handler("image_batch", _image_batch_handler, user_limit=1)


@router.post("/columns")
async def table_columns(
    file: UploadFile = File(...),
    current_user: dict = require_auth(),
):
    """解析上传表格的表头列名（用于前端字段映射下拉，不落盘）。"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "请上传 .xlsx / .csv 文件")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "文件超过 20MB")
    ext = os.path.splitext(file.filename)[1].lower()
    tmp = os.path.join(BATCH_DIR, f"cols_{uuid.uuid4().hex[:8]}{ext}")
    try:
        with open(tmp, "wb") as f:
            f.write(content)
        columns = _parse_header(tmp, file.filename)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    return {"filename": file.filename, "columns": [c for c in columns if c]}


@router.post("/batch")
async def batch_generate(
    template_id: str = Form(...),
    file: UploadFile = File(...),
    field_map: str = Form("{}", description='JSON：{"列名": "图层key"}'),
    batch_name: str = Form(""),
    current_user: dict = require_auth(),
):
    """Excel/CSV 批量套版生成（异步任务，完成后可下载 zip）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    uid = current_user.get("user_id", "") if isinstance(current_user, dict) else ""
    role = current_user.get("role", "") if isinstance(current_user, dict) else ""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "请上传 .xlsx / .csv 文件")
    try:
        fmap = json.loads(field_map or "{}")
        if not isinstance(fmap, dict):
            raise HTTPException(400, "field_map 必须是 JSON 对象")
    except json.JSONDecodeError as e:
        raise HTTPException(400, "服务异常，请稍后重试") from e

    # 预检：模板存在 + 当前用户有权限（收费模板）
    template = load_template(template_id)
    check_render_access(user, template)

    # 保存上传文件（worker 从磁盘读取）
    ext = os.path.splitext(file.filename)[1].lower()
    fpath = os.path.join(BATCH_DIR, f"input_{uuid.uuid4().hex[:8]}{ext}")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "文件超过 20MB")
    with open(fpath, "wb") as f:
        f.write(content)

    task = create_task(
        "image_batch",
        {
            "template_id": template_id,
            "filepath": fpath,
            "filename": file.filename,
            "field_map": fmap,
            "batch_name": batch_name or template.get("name", "批量生成"),
        },
        username=user,
        user_id=uid,
        role=role,
    )
    return {
        "task_id": task["id"],
        "status": "pending",
        "message": f"批量生成任务已提交（共需读取表格行数），可在任务中心查看进度",
        "task": task,
    }


@router.get("/batch/{task_id}/download")
async def batch_download(task_id: str, current_user: dict = require_auth()):
    """下载批量生成结果 zip（校验任务归属）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    from task_queue import get_task

    task = get_task(task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    if task.get("created_by") != user and current_user.get("role") != "admin":
        raise HTTPException(403, "无权下载该任务")
    result = task.get("result") or {}
    zip_path = result.get("zip_path") or ""
    if not zip_path or not os.path.exists(zip_path):
        raise HTTPException(404, "批量结果尚未生成（任务可能仍在执行或已失败）")
    name = os.path.basename(zip_path)
    return FileResponse(zip_path, media_type="application/zip", filename=name)
