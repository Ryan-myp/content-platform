#!/usr/bin/env python3



from typing import Any, Optional, Union, List, Dict, Tuple, Callable, Set, TypeVar, Generic, Iterator, Sequence, Mapping, Iterable, Awaitable, Coroutine, Type
from dataclasses import dataclass, field
from enum import Enum, auto
from datetime import datetime
import asyncio
from typing import Any, Optional, Union, List, Dict, Tuple, Callable, Set, TypeVar, Generic, Iterator, Sequence, Mapping
from dataclasses import dataclass, field
from enum import Enum, auto
from datetime import datetime
"""短剧工厂模块 - LLM 剧本分镜 + CosyVoice 配音 + ffmpeg 视频组装（本地管线）。

流水线：主题 → LLM 剧本（分幕/分镜/台词/旁白）→ 每镜 CosyVoice 配音 →
PIL 镜头背景图 → ffmpeg 逐镜合成 → 拼接 + 字幕烧录 → mp4 产物。
镜头素材当前为本地生成（渐变+文案背景图），后续可平滑替换为数字人口播/云 API 视频素材。
"""

import asyncio
import hashlib
import io
import json
import logging
import math
import os
import re
import shutil
import subprocess
import tempfile
import time
import uuid
from collections.abc import Callable
from pathlib import Path

from fastapi import APIRouter, Form, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse

from common.artifacts import save_artifact
from common.auth import require_auth
from common.config import AGNES_API_BASE, AGNES_API_KEY, load_config, resolve_api_key, resolve_api_base
from common.db import get_db
from common.llm import call_llm_async
from pydantic import BaseModel, Field
from task_queue import create_task, register_handler

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/drama", tags=["短剧工厂"])

load_config()

DRAMA_DIR = Path(__file__).parent / "drama_factory"
DRAMA_DIR.mkdir(parents=True, exist_ok=True)

# ── 素材库（v13.25 开源化：Pexels 实时素材 → 本地素材目录 → 渐变卡片兜底）──
# Pexels 免费 key 注册：https://www.pexels.com/api/ （填入 backend/.env 的 PEXELS_API_KEY）
PEXELS_API_KEY = os.environ.get("PEXELS_API_KEY", "").strip()
MATERIALS_DIR = DRAMA_DIR / "materials"  # 用户本地素材：*关键词*.mp4/jpg（无 key 时兜底）
CACHE_DIR = DRAMA_DIR / "cache"  # Pexels 下载缓存（按 URL 哈希去重）
MUSIC_DIR = DRAMA_DIR / "music"  # 背景音乐目录（*.mp3/wav，可选）
for _d in (MATERIALS_DIR, CACHE_DIR, MUSIC_DIR):
    _d.mkdir(parents=True, exist_ok=True)
from common.helpers import _aggregate_compute_results, _execute_common_step, _execute_compute_step, _execute_single_step, _execute_step, _finalize_common_operation, _finalize_results, _finalize_step_results, _initialize_compute_context, _prepare_common_context, _prepare_context, _prepare_step_context, _notify_progress



def _ffmpeg_bin() -> str:
    """优先使用 imageio-ffmpeg 自带二进制（含 libass，支持 subtitles 滤镜烧录）。"""
    try:
        import imageio_ffmpeg

        return imageio_ffmpeg.get_ffmpeg_exe()
    except Exception:
        return "/usr/local/bin/ffmpeg"


FFMPEG_BIN = _ffmpeg_bin()
FFPROBE_BIN = "/usr/local/bin/ffprobe"
FPS = 25

_VIDEO_EXTS = (".mp4", ".mov", ".webm")
_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".webp")
_MUSIC_EXTS = (".mp3", ".wav", ".m4a", ".aac")


def _resolve_pexels_key() -> str:
    """解析当前请求应使用的 Pexels key：用户级优先（个人中心配置），回退全局环境变量。"""
    try:
        from common.relay_context import get_relay_context

        ctx = get_relay_context()
        if ctx and ctx.get("pexels_key"):
            return ctx["pexels_key"]
    except Exception:
        pass
    return PEXELS_API_KEY


def _pexels_search_video(query: str) -> str | None:
    """Pexels API 按关键词搜索竖屏视频素材，返回合适的 mp4 直链（失败/无 key 返回 None）。

    v13.25 借鉴 MoneyPrinterTurbo 素材管线：LLM 输出每镜搜索关键词 → 实时拉取真实素材。
    v13.29 相关性增强：per_page 5→15、时长过滤 8-40s（利于单镜循环）、
    按 (关键词+日期) 哈希轮换 top 5 候选（同词每天换一批，避免缓存死锁同画面）。
    v1.0.36：key 支持用户级配置（个人中心填写），优先用用户自己的 key。
    """
    _key = _resolve_pexels_key()
    if not _key:
        return None
    try:
        import requests

        r = requests.get(
            "https://api.pexels.com/videos/search",
            params={"query": query, "orientation": "portrait", "per_page": 15},
            headers={"Authorization": _key},
            timeout=10,
        )
        if r.status_code != 200:
            logger.warning(f"Pexels 搜索失败 HTTP {r.status_code}: {query}")
            return None
        links = []
        for v in r.json().get("videos") or []:
            try:
                dur = float(v.get("duration") or 0)
            except (TypeError, ValueError):
                dur = 0.0
            if not (8 <= dur <= 40):  # 太短循环突兀 / 太长超出单镜
                continue
            files = [
                f for f in (v.get("video_files") or [])
                if f.get("file_type") == "video/mp4" and f.get("link") and f.get("width") and f.get("height")
            ]
            pool = [f for f in files if f["height"] >= f["width"]] or files
            for f in pool:
                w = f.get("width") or 0
                if w:
                    links.append((f["link"], w))
        if not links:
            return None
        # 720-1920 宽优先（清晰度与体积平衡），不足时任意宽度兜底
        ok = [lnk for lnk, w in links if 720 <= w <= 1920]
        pool = ok or [lnk for lnk, _ in links]
        picks = pool[:5]
        idx = int(hashlib.sha256(f"{query}|{time.strftime('%Y-%m-%d')}".encode()).hexdigest(), 16) % len(picks)
        return picks[idx]
    except Exception as e:
        logger.warning(f"Pexels 素材搜索异常: {e}")
    return None


def _download_material(url: str, dest: Path) -> bool:
    """流式下载素材到缓存（.part 临时文件 + 原子重命名，失败清理）。"""
    tmp = dest.with_name(dest.name + ".part")
    try:
        import requests

        with requests.get(url, stream=True, timeout=60) as r:
            if r.status_code != 200:
                logger.warning(f"素材下载失败 HTTP {r.status_code}: {url[:80]}")
                return False
            with open(tmp, "wb") as f:
                for chunk in r.iter_content(1 << 16):
                    f.write(chunk)
        tmp.rename(dest)
        return True
    except Exception as e:
        logger.warning(f"素材下载异常: {e}")
        tmp.unlink(missing_ok=True)
        return False


def _find_local_material(query: str) -> Path | None:
    """本地素材目录模糊匹配：文件名含关键词（中文/英文均可）的视频或图片。"""
    q = (query or "").strip()
    if not q:
        return None
    for ext in _VIDEO_EXTS + _IMAGE_EXTS:
        for p in MATERIALS_DIR.rglob(f"*{q}*.{ext.lstrip('.')}"):
            if p.is_file():
                return p
    return None


def _fetch_material(query: str) -> tuple[Path | None, str]:
    """获取镜头素材：Pexels 搜索下载（带 URL 哈希缓存）→ 本地素材目录。返回 (路径, 类型 video/image)。"""
    q = (query or "").strip()
    if not q:
        return None, ""
    url = _pexels_search_video(q)
    if url:
        ext = Path(url.split("?")[0]).suffix.lower() or ".mp4"
        cache = CACHE_DIR / f"{hashlib.sha256(f'{q}|{url}'.encode()).hexdigest()[:16]}{ext}"
        if cache.exists() and cache.stat().st_size > 4096:
            return cache, ("video" if ext in _VIDEO_EXTS else "image")
        if _download_material(url, cache):
            kind = "video" if cache.suffix.lower() in _VIDEO_EXTS else "image"
            return cache, kind
    local = _find_local_material(q)
    if local:
        return local, ("video" if local.suffix.lower() in _VIDEO_EXTS else "image")
    return None, ""

_SCRIPT_SYSTEM = """你是资深短剧编剧。把用户主题扩写成一部节奏紧凑的竖屏短剧脚本。
要求：
1. 输出严格的 JSON（不要 markdown 代码块，不要多余文字）
2. 结构：{"title": "剧名", "characters": [{"id": "lin", "name": "林小满", "gender": "女", "age": "24岁", "appearance": "黑色长发齐刘海，圆脸大眼睛", "outfit": "白色连衣裙配红色围巾", "search": "young chinese woman black hair"}], "scenes": [{"id": 1, "chars": ["lin"], "shot": "镜头画面描述", "search": "英文素材关键词", "narrator": "旁白", "dialogue": "角色台词", "emotion": "情绪", "sec": 25}]}
3. 角色一致性（最重要，v13.30）：先定义全剧 1-3 个主要角色 characters，每个角色的性别/年龄/发型/发色/服装在整部剧固定不变；每镜 chars 列出场角色 id（1-2 个为宜）；每个角色首次出场安排单人镜定妆，后续出场必须沿用该外貌服装；shot 必须点名出场角色并沿用其外貌与服装，禁止临时换外貌或引入新主角形象；search 以该镜主角的英文特征词开头（如 young chinese woman black hair）
4. 场次与目标时长匹配（v13.27 长剧）：场次数 ≈ 目标秒数 ÷ 每场 25-30 秒，向上取整（如 300 秒 → 10-12 场；600 秒 → 20-24 场；≤2 分钟不少于 4 场，最多 28 场）；每场 sec 15-40 秒（单镜保底时长，配音不足时画面素材循环补足）
5. shot 必须是"画面里能拍到的东西"（具体场景/道具/人物动作/天气/光线），禁止抽象概念；search 给出 2-4 个与 shot 完全对应的英文关键词（如 night city rain neon），禁止抽象词（dream/hope/life 等搜不到素材的词）
6. 每场旁白+台词合计约 60-100 字（约 20-30 秒口播量），全场总时长贴近目标时长 ±20%，禁止每场超长台词；台词/旁白必须提到本镜画面里的具体元素（画面有雨才能说"雨"），与 shot 强呼应，禁止写与画面无关的抽象感慨
7. 每镜必须标注景别 shot_size（特写/近景/中景/全景/远景），情绪激烈用特写近景、交代环境用全景、对话用中景；shot 与景别一致
8. 每镜必须标注情绪 emotion（v13.24）：happy 欢快 / sad 悲伤 / angry 激昂愤怒 / gentle 温柔 / serious 严肃 / neutral 自然，台词口吻与情绪一致
9. 剧情有起承转合，结尾留悬念钩子"""


async def _generate_script(theme: str, duration_hint: int, template: dict | None = None) -> dict:
    """LLM 生成剧本（v13.29：worker 与 /script 接口共用；v22：题材模板注入）。

    最多重试 3 次解析（LLM 偶发坏 JSON），返回经时长防御校验的剧本，
    保证接口返回的剧本与最终成片剧本一致（所见即所得）。
    题材模板（drama_templates）注入人设/结构/风格/钩子，让 AI 按爆款套路创作。
    """
    last_err = ""
    tpl_prompt = ""
    if template:
        tpl_prompt = (
            f"\n【题材模板：{template.get('name', '')}】\n"
            f"人设与关系：{template.get('setup', '')}\n"
            f"剧情结构：{template.get('structure', '')}\n"
            f"台词风格：{template.get('style', '')}\n"
            f"开篇钩子：{template.get('hook', '')}"
        )
    for attempt in range(3):
        raw = await call_llm_async(
            _SCRIPT_SYSTEM,
            f"主题：{theme}\n目标时长约 {duration_hint} 秒，场次数与每场秒数按编剧规则匹配。{tpl_prompt}",
            max_tokens=8000,
            temperature=0.85,
            timeout=300,
        )
        try:
            script = _parse_script(raw)
            break
        except (ValueError, json.JSONDecodeError) as e:
            last_err = str(e)
            logger.warning(f"剧本解析失败（第 {attempt + 1} 次）: {e}")
    else:
        raise HTTPException(502, "剧本生成失败，请稍后重试")
    scenes = _enforce_duration(script["scenes"], duration_hint)
    script["scenes"] = scenes
    return script


def _enforce_duration(scenes: list[dict], duration_hint: int) -> list[dict]:
    """v13.28 时长硬校验：场次数 + 台词量 + 单镜 sec 三重防御（LLM 不守编剧规则时的兜底）。

    - 场次上限：每场至少约 20s 起步；
    - 台词预算：口播约 2.5 字/秒（v13.28 情绪镜改 pitch 表达后实测 3-4.8 字/s，
      预算保守取 2.5，保证配音短于单镜 sec 兜底值）；
    - 单镜 sec：超出目标均场值 1.25 倍或不足 0.75 倍的收敛到均场值（配音不足时画面按
      sec 循环补足，sec 失控会直接拉长或缩短成片）。
    """
    max_scenes = min(32, max(4, math.ceil(duration_hint / 20)))
    if len(scenes) > max_scenes:
        scenes = scenes[:max_scenes]
        logger.info(f"[时长防御] 场次截断至 {max_scenes}（目标 {duration_hint}s）")
    budget = max(120, int(duration_hint * 2.5))
    words = sum(
        len(" ".join(x for x in (s.get("narrator"), s.get("dialogue")) if x)) for s in scenes
    )
    if words > budget:
        ratio = budget / words
        for s in scenes:
            for key in ("narrator", "dialogue"):
                v = (s.get(key) or "").strip()
                if not v:
                    continue
                cut = max(1, int(len(v) * ratio))
                s[key] = v[:cut].rstrip() + ("…" if len(v) > cut else "")
        logger.info(f"[时长防御] 台词 {words}→{budget} 字（比例 {ratio:.2f}）")
    if scenes:
        base = max(2, min(45, math.ceil(duration_hint * 0.95 / len(scenes))))
        for s in scenes:
            orig = max(2, int(s.get("sec") or base))
            if orig > base * 1.25 or orig < base * 0.75:
                s["sec"] = base
        words_now = sum(
            len(" ".join(x for x in (s.get("narrator"), s.get("dialogue")) if x)) for s in scenes
        )
        est = sum(max(len(" ".join(x for x in (s.get("narrator"), s.get("dialogue")) if x)) / 1.5, s.get("sec", base)) for s in scenes)
        logger.info(
            f"[时长防御] {len(scenes)} 场 / 台词 {words_now} 字 / 均场 sec {base} / 预估成片约 {est:.0f}s（目标 {duration_hint}s）"
        )
    return scenes


def _parse_characters(data: dict) -> list[dict]:
    """解析剧本角色表（v13.30）：id 规范化去重，生成插画/素材锚定文本。

    每个角色产出 anchor（姓名+性别+年龄+外貌+服装，插画 prompt 锚定用）
    与 search（英文特征词，素材搜索同性别/同特征锚定）。
    """
    chars, seen = [], set()
    for c in data.get("characters") or []:
        if not isinstance(c, dict):
            continue
        cid = re.sub(r"[^a-z0-9]", "", str(c.get("id") or "").lower())
        name = str(c.get("name") or "").strip()[:20]
        if not cid or not name or cid in seen:
            continue
        seen.add(cid)
        gender = str(c.get("gender") or "").strip()[:6]
        age = str(c.get("age") or "").strip()[:10]
        appearance = str(c.get("appearance") or "").strip()[:80]
        outfit = str(c.get("outfit") or "").strip()[:80]
        search = re.sub(r"[\"'\[\]]", "", str(c.get("search") or "").strip())[:60]
        anchor = "，".join(x for x in (name, gender, age, appearance, outfit) if x)
        chars.append({
            "id": cid, "name": name, "gender": gender, "age": age,
            "appearance": appearance, "outfit": outfit, "search": search, "anchor": anchor,
        })
    return chars


def _parse_script(raw: str) -> dict:
    """解析 LLM 剧本 JSON（剥 markdown 代码块/前后噪音，失败抛错）。"""
    text = raw.strip()
    m = re.search(r"```(?:json)?\s*(\{.*\})\s*```", text, re.S)
    if m:
        text = m.group(1)
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError("剧本输出不是 JSON")
    candidate = text[start : end + 1]
    try:
        data = json.loads(candidate)
    except json.JSONDecodeError:
        cleaned = re.sub(r",\s*([\]}])", r"\1", candidate)
        try:
            data = json.loads(cleaned)
        except json.JSONDecodeError:
            data = json.loads(_repair_json_quotes(candidate))
    scenes = data.get("scenes") or []
    if not scenes:
        raise ValueError("剧本没有分镜")
    # v13.30 角色表：先解析，供每镜 chars 白名单校验
    chars = _parse_characters(data)
    char_ids = {c["id"] for c in chars}
    for s in scenes:
        if not s.get("shot") and not s.get("narrator") and not s.get("dialogue"):
            raise ValueError(f"分镜 {s.get('id')} 内容为空")
        if not s.get("sec"):
            s["sec"] = 5
        s["sec"] = max(2, min(45, int(s["sec"])))
        # v13.24 情绪白名单清洗：LLM 可能输出非法/中文情绪标签，非法回落 neutral
        emo = str(s.get("emotion") or "neutral").strip().lower()
        if emo not in ("neutral", "happy", "sad", "angry", "gentle", "serious"):
            emo = {"欢快": "happy", "开心": "happy", "悲伤": "sad", "难过": "sad",
                   "激昂": "angry", "愤怒": "angry", "温柔": "gentle", "严肃": "serious"}.get(emo, "neutral")
        s["emotion"] = emo
        # v1.0.40 景别规范化（漫剧镜头语言）
        _sz = str(s.get("shot_size") or "").strip()
        _sz = _sz.replace("特写", "closeup").replace("近景", "medium").replace("中景", "medium").replace("全景", "wide").replace("远景", "wide")
        s["shot_size"] = _sz if _sz in ("closeup", "medium", "wide") else ""
        # v13.25 素材关键词：search 清洗（限长/去引号），缺失回退 shot 前 30 字符（Pexels 兼容中文）
        search = str(s.get("search") or "").strip()
        search = re.sub(r"[\"'\[\]]", "", search)[:60]
        if not search:
            search = (s.get("shot") or "").strip()[:30]
        s["search"] = search
        # v13.30 出场角色：chars 数组（兼容旧 char 单值），与角色表同一规范化后白名单过滤
        raw_chars = s.get("chars") or s.get("char") or []
        if isinstance(raw_chars, str):
            raw_chars = [raw_chars]
        s["chars"] = [
            re.sub(r"[^a-z0-9]", "", str(x).strip().lower())
            for x in raw_chars
            if re.sub(r"[^a-z0-9]", "", str(x).strip().lower()) in char_ids
        ]
    return {"title": data.get("title") or "未命名短剧", "scenes": scenes, "characters": chars}


def _anchor_search(char: dict | None, search: str) -> str:
    """素材搜索锚定（v13.30）：主角英文特征词前缀 → 素材同性别/同特征（尽力而为）。"""
    q = (search or "").strip()
    if char and char.get("search"):
        return f"{char['search']} {q}".strip()
    return q


PORTRAIT_DIR = DRAMA_DIR / "portraits"  # 角色定妆立绘缓存（漫剧模式：全剧/跨集同脸）

# ── 画风预设（漫剧模式：全剧统一画风，对标红果漫剧）──
ART_STYLES = {
    "guoman": {
        "label": "国漫",
        "desc": "中国国漫风格，精致动漫人物，唯美线条，高饱和色彩",
        "prompt": "中国国漫风格，精致动漫人物立绘，唯美线条，高饱和度色彩，动漫电影质感",
    },
    "hanman": {
        "label": "韩漫",
        "desc": "韩式漫画，时尚精美，柔和光影，精致五官",
        "prompt": "韩国漫画风格，时尚精美动漫人物，柔和光影，精致五官，都市言情漫画质感",
    },
    "3d": {
        "label": "3D 动画",
        "desc": "3D 动画电影渲染，立体人物，柔和光照",
        "prompt": "3D动画电影风格，皮克斯式渲染，立体人物，柔和光照，精致材质",
    },
    "realistic": {
        "label": "写实电影",
        "desc": "真人电影质感，真实光影，细节丰富",
        "prompt": "真人电影风格，写实光影，精致细节，电影剧照质感",
    },
}
DEFAULT_ART_STYLE = "guoman"


def _art_style_prompt(style: str) -> str:
    """画风描述（无效/空回退默认）。"""
    s = (style or "").strip().lower()
    cfg = ART_STYLES.get(s)
    return cfg["prompt"] if cfg else ART_STYLES[DEFAULT_ART_STYLE]["prompt"]


def _portrait_key(cid: str, char: dict, art_style: str = "") -> str:
    """立绘缓存 key：角色 id + 外貌服装 + 画风（外貌/画风变了就重新生成）。"""
    sig = f"{char.get('appearance')}|{char.get('outfit')}|{art_style or DEFAULT_ART_STYLE}"
    return f"{cid}_{hashlib.sha256(sig.encode()).hexdigest()[:8]}"


def _load_char_portrait(cid: str, char: dict, uid: str = "", art_style: str = "",
                       _api_key: str = "", _api_base: str = "") -> bytes | None:
    """读取角色立绘缓存（无则生成）。"""
    key = _portrait_key(cid, char, art_style)
    path = PORTRAIT_DIR / f"{key}.jpg"
    if path.exists() and path.stat().st_size > 4096:
        try:
            return path.read_bytes()
        except Exception:
            pass
    data = _generate_character_portrait(char, uid, art_style, _api_key, _api_base)
    return data


def _save_char_portrait(cid: str, char: dict, data: bytes, art_style: str = "") -> None:
    """持久化角色立绘（跨集复用）。"""
    try:
        PORTRAIT_DIR.mkdir(parents=True, exist_ok=True)
        (PORTRAIT_DIR / f"{_portrait_key(cid, char, art_style)}.jpg").write_bytes(data)
    except Exception:
        pass


def _generate_character_portrait(char: dict, uid: str = "", art_style: str = "", _api_key: str = "", _api_base: str = "") -> bytes | None:
    """生成角色定妆立绘（漫剧模式：全剧同脸同装的核心）。

    根据角色圣经的外貌/服装描述 + 画风预设，生成一张竖屏半身立绘，
    之后每镜都用这张立绘做图生图参考锚定 → 同一角色全剧形象一致。
    _api_key/_api_base：显式传入（to_thread 线程无 ContextVar）。
    """
    if not _api_key:
        _api_key = resolve_api_key()
    if not _api_base:
        _api_base = resolve_api_base()
    if not char or not _api_key:
        return None
    try:
        import base64
        import io
        import requests
        from PIL import Image
        from common.config import IMAGE_MODEL, require_model, resolve_feature_model

        name = char.get("name") or ""
        appearance = char.get("appearance") or ""
        outfit = char.get("outfit") or ""
        gender = char.get("gender") or ""
        prompt = (
            f"{_art_style_prompt(art_style)}。竖屏影视级角色定妆立绘，{gender}角色「{name}」，"
            f"外貌：{appearance}；服装：{outfit}。"
            "半身构图，正面微侧，干净纯色背景，电影级打光，高清细节，"
            "五官清晰，画面中无任何文字，无场景，只有角色。"
        )
        body = {
            "model": require_model(resolve_feature_model(uid, "image", IMAGE_MODEL), "图片"),
            "prompt": prompt,
            "size": "1024x1792",
            "ratio": "9:16",
            "n": 1,
            "extra_body": {"response_format": "url"},
        }
        r = requests.post(
            f"{_api_base}/images/generations",
            headers={"Authorization": f"Bearer {_api_key}", "Content-Type": "application/json"},
            json=body,
            timeout=90,
        )
        r.raise_for_status()
        url = (r.json().get("data") or [{}])[0].get("url")
        if not url:
            return None
        img_resp = requests.get(url, timeout=60)
        if img_resp.status_code != 200:
            return None
        img = Image.open(io.BytesIO(img_resp.content))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"角色立绘生成失败: {api_error_detail(e)}")
        return None


def _generate_scene_image(shot: str, anchors: str = "", refs: list[bytes] | None = None, uid: str = "",
                           art_style: str = "", dialogue: str = "", shot_size: str = "",
                           _api_key: str = "", _api_base: str = "") -> bytes | None:
    """AGNES 文生图/图生图镜头插画（v13.30 角色一致性 + 画风统一）。

    参考图 refs（角色立绘）非空 → 图生图/多图合成锚定角色形象；
    画风 art_style 全剧统一；台词元素 dialogue 注入画面（画面-台词强匹配）。
    统一 scale+crop 到 720x1280；失败返回 None 由调用方回退。
    """
    if not shot or not resolve_api_key():
        return None
    try:
        import base64
        import io
        import requests
        from PIL import Image
        # 函数内取最新配置：config 表运行中修改后无需重启即时生效
        from common.config import IMAGE_MODEL, require_model, resolve_feature_model
        from common.llm import api_error_detail

        # 漫剧模式：画面必须与 shot 描述强一致（红果漫剧标准），角色锚定外貌服装
        dialogue_hint = ""
        if dialogue:
            dialogue_hint = f"人物表情与动作须符合台词语境（{dialogue[:40]}）。"
        size_hint = {
            "closeup": "特写景别：突出人物面部表情与情绪，背景虚化",
            "medium": "中景景别：人物半身入画，兼顾表情与环境",
            "wide": "全景景别：完整交代场景环境，人物融入环境",
        }.get(shot_size, "")
        prompt = (
            f"{_art_style_prompt(art_style)}。竖屏短剧电影分镜插画，竖构图，画面只有场景与人物，"
            f"画面中无任何文字，{dialogue_hint}"
            + (f"{size_hint}。" if size_hint else "")
            + (f"出场的角色必须保持：{anchors}（外貌与服装全剧不变）。" if anchors else "")
            + f"镜头画面：{shot}"
        )
        body = {
            "model": require_model(resolve_feature_model(uid, "image", IMAGE_MODEL), "图片"),
            "prompt": prompt,
            "size": "1024x1792",
            "ratio": "9:16",
            "n": 1,
            "extra_body": {"response_format": "url"},
        }
        if refs:
            # AGNES 支持多图 reference：直接传各角色立绘 + prompt 描述位置区分
            body["image"] = ["data:image/jpeg;base64," + base64.b64encode(r).decode() for r in refs[:3]]
        r = requests.post(
            f"{_api_base}/images/generations",
            headers={"Authorization": f"Bearer {_api_key}", "Content-Type": "application/json"},
            json=body,
            timeout=90,
        )
        r.raise_for_status()
        url = (r.json().get("data") or [{}])[0].get("url")
        if not url:
            return None
        img_resp = requests.get(url, timeout=60)
        if img_resp.status_code != 200:
            return None
        img = Image.open(io.BytesIO(img_resp.content))
        # 等比放大裁剪到 720x1280（与封面/成片一致，无黑边）
        scale = max(720 / img.width, 1280 / img.height)
        img = img.resize((int(img.width * scale), int(img.height * scale)), Image.LANCZOS)
        left, top = (img.width - 720) // 2, (img.height - 1280) // 2
        img = img.crop((left, top, left + 720, top + 1280))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"镜头插画生成失败: {api_error_detail(e)}")
        return None


def _make_scene_card(text: str, idx: int, total: int, title: str, path: str, shot: str = "", uid: str = "") -> bool:
    """PIL 生成镜头背景图：优先 AGNES 插画（shot 画面描述），失败回退渐变海报。

    v13.29 去"大字报"：渐变底不再印整段台词（字幕才是文字载体，避免画面=字幕），
    只保留剧名 + 镜头序号 + 短标题；插画失败静默回退，不阻塞主链路。
    """
    if shot:
        data = _generate_scene_image(shot, uid=uid)
        if data:
            try:
                with open(path, "wb") as f:
                    f.write(data)
                return True
            except OSError:
                pass
    try:
        from PIL import Image, ImageDraw, ImageFont

        w, h = 720, 1280
        grads = [
            ((147, 51, 234), (236, 72, 153)),
            ((59, 130, 246), (16, 185, 129)),
            ((245, 158, 11), (239, 68, 68)),
            ((16, 185, 129), (14, 165, 233)),
            ((99, 102, 241), (217, 70, 239)),
        ]
        c1, c2 = grads[idx % len(grads)]
        img = Image.new("RGB", (w, h))
        draw = ImageDraw.Draw(img)
        for y in range(h):
            t = y / h
            draw.line(
                [(0, y), (w, y)],
                fill=tuple(int(a + (b - a) * t) for a, b in zip(c1, c2, strict=False)),
            )
        try:
            font_big = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 56)
            font_mid = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 36)
        except OSError:
            font_big = font_mid = ImageFont.load_default()
        draw.text((w // 2, 260), title, fill="white", font=font_big, anchor="mm")
        # v13.29 短标题（shot 前 12 字，替代旧版整段台词大字报）
        heading = (shot or text or "").strip()[:12]
        if heading:
            draw.text((w // 2, 520), heading, fill=(255, 255, 255, 230), font=font_mid, anchor="mm")
        draw.text((w // 2, 1060), f"第 {idx + 1} 镜 / 共 {total} 镜", fill=(255, 255, 255, 180), font=font_mid, anchor="mm")
        img.save(path, quality=88)
        return True
    except Exception as e:
        logger.warning(f"镜头背景图生成失败: {e}")
        return False


def _make_intro_card(title: str, art_style: str = "", subtitle: str = "") -> str | None:
    """生成片头标题卡：画风渐变背景 + 剧名大字 + 副标题（红果漫剧开场）。

    返回图片路径（PNG），失败返回 None（不影响主流程）。
    """
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageFilter

        w, h = 720, 1280
        # 按画风选色调
        style = (art_style or DEFAULT_ART_STYLE).strip().lower()
        palettes = {
            "guoman": ((147, 51, 234), (236, 72, 153)),     # 紫粉（国漫）
            "hanman": ((59, 130, 246), (16, 185, 129)),     # 蓝绿（韩漫清新）
            "3d": ((245, 158, 11), (239, 68, 68)),          # 橙红（3D 温暖）
            "realistic": ((30, 30, 30), (120, 120, 120)),   # 灰黑（写实沉稳）
        }
        c1, c2 = palettes.get(style, palettes["guoman"])
        img = Image.new("RGB", (w, h))
        draw = ImageDraw.Draw(img)
        for y in range(h):
            t = y / h
            draw.line([(0, y), (w, y)], fill=tuple(int(a + (b - a) * t) for a, b in zip(c1, c2, strict=False)))
        img = img.filter(ImageFilter.GaussianBlur(radius=4))
        # 装饰：柔和光晕
        glow = Image.new("RGB", (w, h), (0, 0, 0))
        gd = ImageDraw.Draw(glow)
        gd.ellipse([w * 0.15, h * 0.25, w * 0.85, h * 0.75], fill=tuple(int(c * 0.55) for c in c1))
        glow = glow.filter(ImageFilter.GaussianBlur(radius=120))
        img = Image.blend(img, glow, 0.35)
        draw = ImageDraw.Draw(img)
        try:
            font_big = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 84)
            font_sub = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 36)
        except OSError:
            font_big = font_sub = ImageFont.load_default()
        # 剧名居中，带描边阴影
        title = (title or "未命名短剧").strip()
        draw.text((w // 2 + 3, 560 + 3), title, fill=(0, 0, 0, 160), font=font_big, anchor="mm")
        draw.text((w // 2, 560), title, fill="white", font=font_big, anchor="mm")
        if subtitle:
            draw.text((w // 2, 680), subtitle, fill=(255, 255, 255, 200), font=font_sub, anchor="mm")
        path = os.path.join(os.path.dirname(__file__), "drama_factory", f"intro_{int(time.time() * 1000)}.png")
        img.save(path, "PNG")
        return path
    except Exception as e:
        logger.warning(f"片头卡生成失败: {e}")
        return None


def _make_intro_video(img_path: str, bgm_path: str, out_path: str, duration: float = 3.5) -> None:
    """片头视频：标题卡 + BGM 前奏（2s 淡入 + 结尾淡出），Ken Burns 慢推。"""
    total = max(1, int(duration * FPS))
    amp = 0.06
    vf = (
        "scale=1440:2560:force_original_aspect_ratio=increase,crop=1440:2560,"
        f"zoompan=z='1+{amp}*on/{total}':x='iw/2-(iw/zoom/2)':y='ih/2-(ih/zoom/2)':d={total}:s=720x1280:fps={FPS},"
        "fade=t=in:st=0:d=0.4,fade=t=out:st=3.0:d=0.5"
    )
    cmd = [
        FFMPEG_BIN, "-nostdin", "-y",
        "-loop", "1", "-i", img_path,
        "-i", bgm_path,
        "-t", f"{duration:.2f}",
        "-filter_complex",
        f"[1:a]atrim=0:{duration},afade=t=in:st=0:d=0.8,afade=t=out:st={max(0.5, duration - 1.0):.2f}:d=0.8,volume=0.5[a]",
        "-map", "0:v", "-map", "[a]",
        "-vf", vf,
        "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
        "-c:a", "aac", "-b:a", "128k", "-r", str(FPS),
        out_path,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=180)
    if r.returncode != 0 or not os.path.exists(out_path) or os.path.getsize(out_path) < 4096:
        raise RuntimeError("片头视频合成失败: " + r.stderr.decode(errors="replace")[-200:])


def _i2v_motion_prompt(emotion: str, dialogue: str = "") -> str:
    """按情绪生成图生视频动作提示词（人物动态，红果漫剧微动效）。"""
    emo = (emotion or "neutral").strip().lower()
    base = {
        "happy": "gently smiles and nods, eyes brighten, subtle happy expression, natural head movement",
        "sad": "slowly lowers eyes, slight trembling of lips, sad expression, gentle sigh, subtle shoulder movement",
        "angry": "frowns, jaw tightens, intense eyes, slight head shake of disbelief, tense expression",
        "gentle": "softly smiles, eyes warm, gentle nod, calm breathing, tender expression",
        "serious": "slowly narrows eyes, thoughtful expression, subtle chin movement, focused gaze",
        "neutral": "naturally blinks and slightly turns head, subtle breathing, calm expression",
    }.get(emo, "naturally blinks and slightly turns head, subtle breathing, calm expression")
    if dialogue:
        # 台词语境：嘴部微动（口型感）
        base += ", lips move slightly as if speaking, subtle mouth animation"
    return base + ", cinematic, realistic motion, smooth"


def _i2v_scene_clip(img_path: str, prompt: str, out_path: str, uid: str = "", max_wait: int = 240,
                   _api_key: str = "", _api_base: str = "") -> bool:
    """图生视频（i2v）：主图 → AGNES 动态视频 → 裁剪竖屏 → 保存。

    返回是否成功；队列满/超时/失败均返回 False（由调用方回退静态子镜，不阻塞）。
    _api_key/_api_base：显式传入（to_thread 线程拿不到 ContextVar，必须由调用方传入）。
    """
    if not _api_key:
        _api_key = resolve_api_key()
    if not _api_base:
        _api_base = resolve_api_base()
    if not _api_key:
        return False
    try:
        import base64 as _b64
        import requests as _req

        img_b64 = _b64.b64encode(open(img_path, "rb").read()).decode()
        body = {
            "model": "agnes-video-v2.0",
            "prompt": prompt,
            "image": f"data:image/jpeg;base64,{img_b64}",
            "duration": 5,
        }
        resp = _req.post(
            f"{_api_base}/videos",
            headers={"Authorization": f"Bearer {_api_key}", "Content-Type": "application/json"},
            json=body, timeout=60,
        )
        # 队列满：等待重试（最多 3 次×30s），保证动态密度；仍失败才回退静态
        if resp.status_code == 503 and "queue_full" in resp.text:
            for _r in range(3):
                logger.warning(f"i2v 队列满，等待重试（{_r + 1}/3）…")
                time.sleep(30)
                resp = _req.post(
                    f"{_api_base}/videos",
                    headers={"Authorization": f"Bearer {_api_key}", "Content-Type": "application/json"},
                    json=body, timeout=60,
                )
                if resp.status_code == 200:
                    break
            if resp.status_code != 200:
                logger.warning("i2v 队列持续满，回退静态镜头")
                return False
        if resp.status_code != 200:
            logger.warning(f"i2v 提交失败 HTTP {resp.status_code}")
            return False
        vid = (resp.json().get("video_id") or resp.json().get("task_id") or "").strip()
        if not vid:
            return False
        # 轮询
        url = ""
        for _ in range(int(max_wait / 15) + 1):
            time.sleep(15)
            q = _req.get(
                f"{_api_base}/agnesapi", params={"video_id": vid},
                headers={"Authorization": f"Bearer {_api_key}"}, timeout=30,
            )
            try:
                d = q.json()
            except Exception:
                continue
            st = d.get("status")
            if st == "completed":
                url = d.get("output", {}).get("video_url") or d.get("url") or ""
                break
            if st == "failed":
                return False
        if not url:
            return False
        # 下载 + 裁剪竖屏（i2v 输出 1088x832 横屏 → 720x1280 竖屏）
        vresp = _req.get(url, timeout=120)
        if vresp.status_code != 200:
            return False
        tmp_v = out_path + ".src.mp4"
        with open(tmp_v, "wb") as f:
            f.write(vresp.content)
        cmd = [
            FFMPEG_BIN, "-nostdin", "-y", "-i", tmp_v,
            "-vf", "scale=720:1280:force_original_aspect_ratio=increase,crop=720:1280",
            "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-b:a", "128k", out_path,
        ]
        r = subprocess.run(cmd, capture_output=True, timeout=120)
        try:
            os.remove(tmp_v)
        except Exception:
            pass
        if r.returncode != 0 or not os.path.exists(out_path) or os.path.getsize(out_path) < 10000:
            return False
        return True
    except Exception as e:
        logger.warning(f"i2v 场景镜头失败: {e}")
        return False


def _mix_dyn_audio(video_path: str, audio_path: str, out_path: str) -> bool:
    """动态 i2v 画面 + 子镜配音合成（画面时长取两者短者，配音铺满）。"""
    try:
        vdur = _probe_seconds(video_path)
        adur = _probe_seconds(audio_path)
        dur = max(2.0, min(vdur or 5.0, adur or 5.0))
        cmd = [
            FFMPEG_BIN, "-nostdin", "-y", "-i", video_path, "-i", audio_path,
            "-t", f"{dur:.2f}",
            "-map", "0:v", "-map", "1:a",
            "-c:v", "copy",
            "-c:a", "aac", "-b:a", "128k",
            out_path,
        ]
        r = subprocess.run(cmd, capture_output=True, timeout=120)
        return r.returncode == 0 and os.path.exists(out_path) and os.path.getsize(out_path) > 10000
    except Exception:
        return False


def _probe_seconds(path: str) -> float:
    """ffprobe 读取音/视频时长（秒），失败返回 0。"""
    try:
        out = subprocess.run(
            [FFPROBE_BIN, "-v", "error", "-show_entries", "format=duration", "-of", "csv=p=0", path],
            capture_output=True,
            text=True,
            timeout=20,
        )
        return float(out.stdout.strip() or 0)
    except Exception:
        return 0.0


# 子镜头机位类型（同一主图的不同取景方式 → 视觉上的镜头切换）
_SUB_SHOT_TYPES = (
    "zoom_in",      # 推近：从全景推至近景（聚焦）
    "zoom_out",     # 拉远：从近景拉至全景（交代环境）
    "pan_left",     # 左移：取景窗从左向右扫（环境过渡）
    "pan_right",    # 右移：取景窗从右向左扫
    "tilt_up",      # 上移：从下往上（强调高度/气场）
    "tilt_down",    # 下移：从上往下（压迫/揭示）
    "close_zoom",   # 特写放大：聚焦面部/细节（情绪戏）
    "slow_push",    # 缓慢推进：情绪沉淀（温柔/悲伤）
)


def _shot_sequence(emotion: str, n: int, scene_idx: int) -> list[str]:
    """场次内机位递进序列（专业剪辑：景别渐进产生情绪递进）。

    结构：交代/进入（全景移动）→ 中景对话 → 特写聚焦 → 收尾（按情绪）
    - happy/angry：收尾快速推近（情绪高涨）
    - sad/gentle：收尾缓慢拉远（留白余韵）
    - serious：收尾缓慢推进（悬而未决）
    """
    emo = (emotion or "neutral").strip().lower()
    # 基础递进模板
    base = []
    if n >= 4:
        base += ["zoom_out", "pan_left", "pan_right", "zoom_in"]
    elif n == 3:
        base = ["zoom_out", "zoom_in", "slow_push"]
    elif n == 2:
        base = ["zoom_out", "zoom_in"]
    else:
        base = ["slow_push"]
    # 补足到 n：中间插入 close_zoom/tilt（情绪戏）
    extra = ["close_zoom", "tilt_up", "tilt_down", "pan_left", "pan_right"]
    while len(base) < n:
        base.insert(len(base) - 1, extra[(scene_idx + len(base)) % len(extra)])
    base = base[:n]
    # 收尾机位按情绪
    if len(base) >= 2:
        if emo in ("happy", "angry"):
            base[-1] = "close_zoom" if n >= 4 else "zoom_in"
        elif emo in ("sad", "gentle"):
            base[-1] = "zoom_out"
        elif emo == "serious":
            base[-1] = "slow_push"
    return base


def _scene_video(img_path: str, audio_path: str, out_path: str, duration: float,
                 motion: str = "zoom_in", fade_in: bool = True, fade_out: bool = True,
                 win: tuple = (0, 0, 1.0, 1.0)) -> None:
    """单镜合成：背景图 + 配音 → mp4 片段（Ken Burns 运镜 + 可选首尾淡入淡出）。

    v13.31 插画镜流畅度：zoompan 运镜（motion 交替推近/拉远/横摇，静态图动起来）；
    注意 zoompan 必须用 on（输出帧计数）——in 是输入帧计数，按需求值只拉 1 个
    输入帧时全部输出帧共享 in=0，画面静止（素材模式旧代码踩过此坑）；
    fade_in/fade_out 按镜序控制（首镜淡入、末镜淡出、中间镜硬切），消除镜间黑场闪烁。
    无声段 apad 补静音到 -t 目标时长（v13.28 移除 -shortest，sec 画面保底生效）。

    v1.0.41 子镜头：win=(nx,ny,nw,nh) 指定主图取景窗口（0-1 归一化），
    不同子镜取不同区域+不同运动方向 → 同一主图也能切出多个镜头（视觉连续感）。
    """
    total = max(1, int(duration * FPS))
    # 取景窗口（归一化）：默认全图；子镜头给局部区域
    nx, ny, nw, nh = win
    nw = max(0.3, min(1.0, nw)); nh = max(0.3, min(1.0, nh))
    nx = max(0.0, min(1.0 - nw, nx)); ny = max(0.0, min(1.0 - nh, ny))
    # 短镜（<8s）幅度加大、长镜放缓：保证 2-4s 子镜头也有可见运镜
    amp = 0.14 if total < 200 else (0.10 if total < 400 else 0.06)
    # 2x 放大防抖基础缩放：先裁取景窗口再放大
    vf = "scale=1440:2560:force_original_aspect_ratio=increase,crop=1440:2560"
    # 应用取景窗口（crop 到窗口区域，窗口比例保持 9:16 输出）
    win_w = int(1440 * nw); win_h = int(2560 * nh)
    win_x = int(1440 * nx); win_y = int(2560 * ny)
    # 窗口先切出（保持 9:16 比例：以窗口中心为基准放大到满幅）
    vf += f",crop={win_w}:{win_h}:{win_x}:{win_y}"
    vf += ",scale=1440:2560:force_original_aspect_ratio=increase,crop=1440:2560"
    if motion and motion != "still":
        # 各机位：zoom 表达式 + 取景窗中心漂移方向
        if motion == "zoom_out":
            zexpr = f"{1 + amp}-{amp}*on/{total}"
            sx = sy = ""
        elif motion == "pan_left":
            zexpr = f"1+{amp * 0.3}*on/{total}"
            sx = f"-(iw*{amp * 0.5})*on/{total}"
            sy = ""
        elif motion == "pan_right":
            zexpr = f"1+{amp * 0.3}*on/{total}"
            sx = f"+(iw*{amp * 0.5})*on/{total}"
            sy = ""
        elif motion == "tilt_up":
            zexpr = f"1+{amp * 0.3}*on/{total}"
            sx = ""
            sy = f"-(ih*{amp * 0.5})*on/{total}"
        elif motion == "tilt_down":
            zexpr = f"1+{amp * 0.3}*on/{total}"
            sx = ""
            sy = f"+(ih*{amp * 0.5})*on/{total}"
        elif motion == "close_zoom":
            zexpr = f"{1 + amp * 1.4}-{amp * 0.8}*on/{total}"  # 从近景再推近
            sx = sy = ""
        elif motion == "slow_push":
            zexpr = f"1+{amp * 0.6}*on/{total}"
            sx = sy = ""
        else:  # zoom_in
            zexpr = f"1+{amp}*on/{total}"
            sx = sy = ""
        vf += (
            f",zoompan=z='{zexpr}':x='iw/2-(iw/zoom/2){sx}':y='ih/2-(ih/zoom/2){sy}':d={total}:s=720x1280:fps={FPS}"
        )
    else:
        # 静止子镜也做微推（避免死画面）
        vf += f",zoompan=z='1+{amp * 0.3}*on/{total}':x='iw/2-(iw/zoom/2)':y='ih/2-(ih/zoom/2)':d={total}:s=720x1280:fps={FPS}"
    if fade_in:
        vf += ",fade=t=in:st=0:d=0.15"
    if fade_out:
        vf += f",fade=t=out:st={max(0.15, duration - 0.15):.2f}:d=0.15"
    cmd = [
        FFMPEG_BIN, "-nostdin", "-y",
        "-loop", "1", "-i", img_path,
        "-i", audio_path,
        "-t", f"{duration:.2f}",
        "-vf", vf,
        "-c:v", "libx264", "-preset", "fast", "-pix_fmt", "yuv420p",
        "-af", "apad", "-c:a", "aac", "-b:a", "128k",
        "-r", str(FPS),
        out_path,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=180)
    if r.returncode != 0 or not os.path.exists(out_path) or os.path.getsize(out_path) < 4096:
        raise RuntimeError("单镜合成失败: " + r.stderr.decode(errors="replace")[-200:])


_SCENE_MOTIONS = ("zoom_in", "zoom_out", "pan_in", "pan_out")

# 情绪 → 运镜（漫剧模式：情绪不同，镜头语言不同，红果漫剧标准）
_EMOTION_MOTION = {
    "happy": "zoom_in",       # 欢快：推近聚焦
    "gentle": "pan_out",      # 温柔：缓慢横移
    "sad": "zoom_out",        # 悲伤：拉远留白
    "angry": "zoom_in",       # 激昂：快速推近（幅度大）
    "serious": "pan_in",      # 严肃：缓慢推近
    "neutral": "zoom_in",
}


def _motion_for(sc: dict, idx: int) -> str:
    """按情绪选运镜（fallback 循环交替，保证镜间不单调）。"""
    emo = (sc.get("emotion") or "neutral").strip().lower()
    m = _EMOTION_MOTION.get(emo)
    if m:
        return m
    return _SCENE_MOTIONS[idx % len(_SCENE_MOTIONS)]


def _split_audio_segments(audio_path: str, seg_count: int) -> list[str]:
    """把配音音频切成 seg_count 段（每段对应一个子镜头），返回段文件路径列表。

    用 ffmpeg asplit+atrim 一次切出所有段，避免多次读文件；段间 30ms 微交叉
    防爆音（afade in/out 各 15ms）。
    """
    if seg_count <= 1:
        return [audio_path]
    segs = []
    dur = _probe_seconds(audio_path)
    if dur <= 0:
        return [audio_path]
    seg_len = dur / seg_count
    tmpdir = os.path.dirname(audio_path)
    # 简单实现：逐段切（段数不多，逐次调用更稳）；afade 只做入淡防止段首爆音
    out_files = []
    for i in range(seg_count):
        st = i * seg_len
        o = os.path.join(tmpdir, f"subseg_{i:03d}.mp3")
        r = subprocess.run(
            [FFMPEG_BIN, "-nostdin", "-y", "-ss", f"{st:.3f}", "-t", f"{seg_len:.3f}",
             "-i", audio_path, "-af", "afade=t=in:st=0:d=0.015",
             "-c:a", "libmp3lame", "-b:a", "128k", o],
            capture_output=True, timeout=60,
        )
        if r.returncode == 0 and os.path.exists(o) and os.path.getsize(o) > 1024:
            out_files.append(o)
        else:
            logger.warning(f"音频切段 {i} 失败: {(r.stderr or b'')[:120] if r else ''}")
    return out_files if len(out_files) >= 2 else [audio_path]


def _concat_sub_shots(seg_paths: list[str], out_path: str) -> None:
    """拼接子镜头片段（逐段 concat，编码一致可直接拼接）。"""
    list_file = out_path + ".txt"
    with open(list_file, "w", encoding="utf-8") as f:
        for p in seg_paths:
            f.write("file '" + p + "'\n")
    r = subprocess.run(
        [FFMPEG_BIN, "-nostdin", "-y", "-f", "concat", "-safe", "0", "-i", list_file, "-c", "copy", out_path],
        capture_output=True, timeout=300,
    )
    os.remove(list_file)
    if r.returncode != 0 or not os.path.exists(out_path):
        raise RuntimeError("子镜头拼接失败: " + r.stderr.decode(errors="replace")[-200:])


def _material_scene_video(query: str, audio_path: str, out_path: str, duration: float,
                          fade_in: bool = True, fade_out: bool = True) -> bool:
    """素材镜头合成：Pexels/本地真实素材（视频 cover 裁剪或图片 Ken Burns 推近）+ 配音。

    v13.25 核心升级（借鉴 MoneyPrinterTurbo 素材管线）：告别纯渐变卡片，镜头画面改为
    真实视频/图片素材；素材不可用时返回 False，由上层逐镜卡片兜底。
    """
    material, kind = _fetch_material(query)
    if material is None:
        return False
    try:
        vf_base = "scale=720:1280:force_original_aspect_ratio=increase,crop=720:1280"
        if fade_in:
            vf_base += ",fade=t=in:st=0:d=0.25"
        if fade_out:
            vf_base += f",fade=t=out:st={max(0.25, duration - 0.25):.2f}:d=0.25"
        if kind == "video":
            # cover 裁剪无黑边 + 短素材循环补足时长 + 丢弃素材原音、混入配音
            # v13.28 移除 -shortest 改 apad：短配音时画面按 sec 循环补足到目标时长
            cmd = [
                FFMPEG_BIN, "-nostdin", "-y",
                "-stream_loop", "-1", "-i", str(material),
                "-i", audio_path,
                "-t", f"{duration:.2f}",
                "-map", "0:v:0", "-map", "1:a:0",
                "-vf", vf_base,
                "-c:v", "libx264", "-preset", "fast", "-pix_fmt", "yuv420p",
                "-af", "apad", "-c:a", "aac", "-b:a", "128k",
                "-r", str(FPS),
                out_path,
            ]
        else:
            # 图片 Ken Burns：2x 放大防抖 + zoompan 缓慢推近 + 配音
            # v13.31 修复：zoompan 必须用 on（输出帧计数），in 是输入帧计数会导致画面静止
            total = max(1, int(duration * FPS))
            amp = 0.10 if total >= 250 else 0.06
            cmd = [
                FFMPEG_BIN, "-nostdin", "-y",
                "-loop", "1", "-i", str(material),
                "-i", audio_path,
                "-t", f"{duration:.2f}",
                "-map", "0:v:0", "-map", "1:a:0",
                "-vf", (
                    "scale=1440:2560:force_original_aspect_ratio=increase,crop=1440:2560,"
                    f"zoompan=z='1+{amp}*on/{total}':x='iw/2-(iw/zoom/2)':y='ih/2-(ih/zoom/2)':d={total}:s=720x1280:fps={FPS},"
                    + vf_base.split(",", 1)[1]  # 复用 fade 段（去掉 cover 裁剪前缀）
                ),
                "-c:v", "libx264", "-preset", "fast", "-pix_fmt", "yuv420p",
                "-af", "apad", "-c:a", "aac", "-b:a", "128k",
                "-r", str(FPS),
                out_path,
            ]
        r = subprocess.run(cmd, capture_output=True, timeout=300)
        if r.returncode != 0 or not os.path.exists(out_path) or os.path.getsize(out_path) < 4096:
            logger.warning(f"素材镜头合成失败: {r.stderr.decode(errors='replace')[-200:]}")
            return False
        return True
    except Exception as e:
        logger.warning(f"素材镜头合成异常: {e}")
        return False


def _concat_videos(clip_paths: list[str], out_path: str, scene_bounds: list[int] | None = None) -> None:
    """拼接镜头片段：子镜头间硬切（快节奏），场次间交叉淡化（大段落转场）。

    scene_bounds: 场次起始 clip 下标列表（如 [0, 5, 11] 表示第 0/5/11 个 clip 是新场次）。
    场次边界用 xfade 交叉淡化（0.35s），其余硬切（concat demuxer，快速）。
    """
    if not scene_bounds or len(scene_bounds) < 2 or len(clip_paths) < 2:
        list_file = out_path + ".txt"
        with open(list_file, "w", encoding="utf-8") as f:
            for pp in clip_paths:
                f.write("file '" + pp + "'\n")
        r = subprocess.run(
            [FFMPEG_BIN, "-nostdin", "-y", "-f", "concat", "-safe", "0", "-i", list_file, "-c", "copy", out_path],
            capture_output=True,
            timeout=300,
        )
        os.remove(list_file)
        if r.returncode != 0 or not os.path.exists(out_path):
            raise RuntimeError("片段拼接失败: " + r.stderr.decode(errors="replace")[-200:])
        return
    group_starts = set(scene_bounds)
    groups: list[list[str]] = []
    cur: list[str] = []
    for i, c in enumerate(clip_paths):
        if i in group_starts and cur:
            groups.append(cur)
            cur = [c]
        else:
            cur.append(c)
    if cur:
        groups.append(cur)
    tmp_group = []
    for gi, g in enumerate(groups):
        if len(g) == 1:
            tmp_group.append(g[0])
        else:
            gp = os.path.join(os.path.dirname(out_path), "grp_%d.mp4" % gi)
            _concat_videos(g, gp)
            tmp_group.append(gp)
    if len(tmp_group) == 1:
        _concat_videos(tmp_group, out_path)
        return
    inputs = []
    for pp in tmp_group:
        inputs += ["-i", pp]
    group_durs = [_probe_seconds(pp) for pp in tmp_group]
    offsets = []
    acc = (group_durs[0] - 0.35) if group_durs else 0.0
    for gi in range(1, len(tmp_group)):
        offsets.append(max(0.0, acc))
        if gi < len(group_durs):
            acc = acc + (group_durs[gi] - 0.35)
    f = ""
    for gi in range(1, len(tmp_group)):
        off = offsets[gi - 1] if gi - 1 < len(offsets) else 0.0
        if gi == 1:
            f += "[0:v][1:v]xfade=transition=fade:duration=0.35:offset=%.3f[v%d]" % (off, gi)
        else:
            f += "[v%d][%d:v]xfade=transition=fade:duration=0.35:offset=%.3f[v%d]" % (gi - 1, gi, off, gi)
    for gi in range(1, len(tmp_group)):
        if gi == 1:
            f += "[0:a][1:a]acrossfade=d=0.35[a%d]" % gi
        else:
            f += "[a%d][%d:a]acrossfade=d=0.35[a%d]" % (gi - 1, gi, gi)
    last = len(tmp_group) - 1
    r = subprocess.run(
        [FFMPEG_BIN, "-nostdin", "-y"] + inputs + ["-filter_complex", f,
         "-map", "[v%d]" % last, "-map", "[a%d]" % last,
         "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
         "-c:a", "aac", "-b:a", "128k", out_path],
        capture_output=True, timeout=600,
    )
    for gp in tmp_group:
        if gp.startswith(os.path.dirname(out_path) + os.sep + "grp_"):
            try:
                os.remove(gp)
            except Exception:
                pass
    if r.returncode != 0 or not os.path.exists(out_path):
        _concat_videos(clip_paths, out_path)
        return

def _pick_bgm() -> str | None:
    """选一首背景音乐：优先 drama_factory/music（用户自备），
    其次复用音乐工厂生成的作品（music_factory/*.mp3）——零成本氛围感。
    目录为空/无音频返回 None。
    """
    import random

    candidates: list[str] = []
    if MUSIC_DIR.exists():
        candidates += [
            str(p) for p in MUSIC_DIR.iterdir()
            if p.is_file() and p.suffix.lower() in _MUSIC_EXTS
        ]
    mf_dir = DRAMA_DIR.parent / "music_factory"
    if mf_dir.exists():
        candidates += [
            str(p) for p in mf_dir.glob("*.mp3")
            if p.is_file() and p.stat().st_size > 50_000
        ]
    if not candidates:
        return None
    return random.choice(candidates)


def _burn_subtitles(video_path: str, srt_path: str, out_path: str, bgm_path: str | None = None, margin_v: int = 24) -> None:
    """字幕烧录（subtitles 滤镜）+ 背景音乐混音（v13.25：合并一次 re-encode）。

    BGM 音量 12%（配音优先）+ 首尾 2s 淡入淡出；无 BGM 时保持原逻辑。
    margin_v: 字幕底部安全区边距（红果竖屏规范：字幕不贴底、不挡脸）。
    """
    esc = srt_path.replace(":", "\\:").replace("'", "\\'")
    subtitle_vf = f"subtitles='{esc}':force_style='FontName=PingFang SC,FontSize=14,PrimaryColour=&H00FFFFFF,OutlineColour=&H80000000,Outline=1,Shadow=0,MarginV={int(margin_v)}'"
    if bgm_path and os.path.exists(bgm_path):
        total = max(_probe_seconds(video_path), 1.0)
        fade_st = max(0.0, total - 2.0)
        try:
            cmd = [
                FFMPEG_BIN, "-nostdin", "-y", "-i", video_path, "-i", bgm_path,
                "-filter_complex",
                f"[1:a]atrim=0:{total:.2f},aresample=44100,volume=0.12,"
                f"afade=t=in:st=0:d=2,afade=t=out:st={fade_st:.2f}:d=2[bgm];"
                f"[0:a][bgm]amix=inputs=2:duration=first:dropout_transition=0:normalize=0[a]",
                "-map", "0:v", "-map", "[a]",
                "-vf", subtitle_vf,
                "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
                "-c:a", "aac", "-b:a", "128k",
                out_path,
            ]
            r = subprocess.run(cmd, capture_output=True, timeout=600)
            if r.returncode == 0 and os.path.exists(out_path):
                return
            logger.warning(f"BGM 混音字幕失败，回退无 BGM: {(r.stderr or b'')[-120:]}")
        except Exception as e:
            logger.warning(f"BGM 混音异常，回退无 BGM: {e}")
    # 无 BGM 字幕烧录（含 BGM 失败回退）
    cmd = [
        FFMPEG_BIN, "-nostdin", "-y", "-i", video_path,
        "-vf", subtitle_vf,
        "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
        "-c:a", "copy",
        out_path,
    ]
    r = subprocess.run(cmd, capture_output=True, timeout=600)
    if r.returncode != 0 or not os.path.exists(out_path):
        raise RuntimeError("字幕烧录失败: " + r.stderr.decode(errors="replace")[-200:])


def _qc_check(final_path: str, srt_path: str | None = None, min_duration: float = 10.0) -> dict:
    """短剧成片 QC 质量门：时长/黑边/音频/字幕/画面完整性。

    返回 {"ok": bool, "findings": [str]}；不阻塞成片，仅标记质量问题。
    """
    findings = []
    path = str(final_path)
    if not os.path.exists(path) or os.path.getsize(path) < 50_000:
        return {"ok": False, "findings": ["成片文件缺失或过小"]}
    # 时长
    dur = _probe_seconds(path)
    if dur < min_duration:
        findings.append(f"成片时长过短（{dur:.1f}s < {min_duration:.0f}s）")
    # 音轨存在（ffprobe 探测更稳：动态锚/i2v 音轨可能让 ffmpeg 解码探测误报）
    try:
        r = subprocess.run(
            [FFPROBE_BIN, "-v", "error", "-select_streams", "a:0",
             "-show_entries", "stream=codec_name", "-of", "csv=p=0", path],
            capture_output=True, text=True, timeout=30,
        )
        if r.returncode != 0 or not (r.stdout or "").strip():
            findings.append("成片缺少音轨")
    except Exception:
        findings.append("音轨检查失败")
    # 字幕文件存在且非空
    if srt_path and os.path.exists(srt_path):
        if os.path.getsize(srt_path) < 30:
            findings.append("字幕文件为空")
    elif srt_path:
        findings.append("字幕文件缺失")
    # 分辨率/黑边：竖屏 720x1280 期望（ffprobe 宽高）
    try:
        r = subprocess.run(
            [FFPROBE_BIN, "-v", "error", "-select_streams", "v:0",
             "-show_entries", "stream=width,height", "-of", "csv=p=0", path],
            capture_output=True, text=True, timeout=30,
        )
        parts = (r.stdout or "").strip().split(",")
        if len(parts) == 2:
            w, h = int(parts[0]), int(parts[1])
            # 竖屏短剧应为 9:16（宽:高 ≈ 0.5625）；检测横屏 16:9（宽:高 ≈ 1.777）
            if h > 0 and w > h and abs(w / h - 16 / 9) < 0.05:
                findings.append(f"疑似横屏（{w}x{h}），竖屏短剧应为 9:16")
            if h > 0 and (w % 2 != 0 or h % 2 != 0):
                findings.append(f"分辨率非偶数（{w}x{h}），部分平台播放异常")
    except Exception:
        pass
    return {"ok": len(findings) == 0, "findings": findings}


def _extract_cover(video_path: str, out_jpg: str) -> None:
    """从首镜视频抽帧生成真 JPG 封面（竖屏 720x1280 无黑边，失败静默）。"""
    try:
        subprocess.run(
            [
                FFMPEG_BIN, "-nostdin", "-y", "-ss", "0.4", "-i", video_path,
                "-frames:v", "1",
                "-vf", "scale=720:1280:force_original_aspect_ratio=increase,crop=720:1280",
                "-q:v", "2", out_jpg,
            ],
            capture_output=True,
            timeout=30,
        )
    except Exception as e:
        logger.warning(f"封面抽帧失败: {e}")


def _make_preview(video_path: str, out_mp4: str) -> None:
    """截取首镜前 6 秒作列表 hover 动态预览（失败静默，前端回退静态封面）。"""
    try:
        subprocess.run(
            [
                FFMPEG_BIN, "-nostdin", "-y", "-ss", "0", "-t", "6", "-i", video_path,
                "-c:v", "libx264", "-preset", "fast", "-c:a", "aac",
                "-movflags", "+faststart", out_mp4,
            ],
            capture_output=True,
            timeout=60,
        )
    except Exception as e:
        logger.warning(f"预览视频生成失败: {e}")


def _srt_ts(seconds: float) -> str:
    """秒数 → SRT 时间戳（HH:MM:SS,mmm）。"""
    ms = int(round(seconds * 1000))
    h, rem = divmod(ms, 3600000)
    m, rem = divmod(rem, 60000)
    s, msec = divmod(rem, 1000)
    return f"{h:02d}:{m:02d}:{s:02d},{msec:03d}"


def _split_sentences(text: str) -> list[str]:
    """台词按句切分（句号/问号/感叹号/省略号；含旁白标记的行单独成句）。"""
    import re as _re

    parts = _re.split(r"(?<=[。！？…])", text or "")
    out = [p.strip() for p in parts if p.strip()]
    # 无标点大段（LLM 偶发）按最长 20 字硬切
    final = []
    for p in out:
        while len(p) > 20:
            final.append(p[:20])
            p = p[20:]
        if p:
            final.append(p)
    return final or ([text.strip()] if text.strip() else [])


def _make_srt(scenes: list[dict], durations: list[float], voice_durs: list[float], out_path: str) -> None:
    """生成 SRT：逐句字幕（红果漫剧标准——字幕跟随说话节奏，与快镜头切换同步）。

    v13.29 字幕时序收敛：显示时长 = min(画面时长, 配音时长 + 0.6s)；
    v1.0.43 逐句化：旁白/台词按句切分，每句按配音时长比例分配时间，
    字幕逐句出现（长句不再一屏到底），观感跟随 2 秒镜头节奏。
    voice_durs 缺省时退化为整镜显示（数字人模式全镜有声）。
    """

    ts = _srt_ts

    lines, cursor = [], 0.0
    idx = 1
    for i, (sc, dur) in enumerate(zip(scenes, durations, strict=False), 1):
        vd = voice_durs[i - 1] if i - 1 < len(voice_durs) else dur
        show = min(max(dur, 1.0), max(float(vd) + 0.6, 1.2)) if vd else max(dur, 1.0)
        start = cursor
        text = " ".join(x for x in (sc.get("narrator"), sc.get("dialogue")) if x)
        sents = _split_sentences(text)
        if len(sents) <= 1:
            lines.append(f"{idx}\n{ts(start)} --> {ts(start + show)}\n{text.strip()}\n")
            idx += 1
        else:
            # 多句：按句长比例分配配音时长（每句至少 0.8s）
            total_chars = max(1, sum(len(s) for s in sents))
            t = start
            for s in sents:
                s_dur = max(0.8, show * len(s) / total_chars)
                lines.append(f"{idx}\n{ts(t)} --> {ts(t + s_dur)}\n{s}\n")
                idx += 1
                t += s_dur
        cursor += max(dur, 1.0)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _tts_scene(text: str, emotion: str = "neutral") -> bytes:
    """单镜配音：复用 voice_factory._tts_one 全降级链路（CosyVoice → edge → 中转站）。

    v13.24 情绪：neutral 保持 CosyVoice「中文女」高质量音色；带情绪镜切 Azure 音色。
    v13.28 起情绪在 voice_factory._tts_edge 内改为 pitch 叠加表达（SSML style 语速黑洞），
    此处仍传 style 别名（cheerful 等）以便 _tts_edge 映射，兼容旧调用链。
    """
    from voice_factory import _tts_one

    if emotion and emotion != "neutral":
        style = {"happy": "cheerful", "sad": "sad", "angry": "angry",
                 "gentle": "gentle", "serious": "serious"}.get(emotion, "")
        if style:
            return _tts_one(text, "zh-CN-XiaoxiaoNeural", 1.05, 0, style)
    return _tts_one(text, "中文女", 1.05)


def _dh_scene_video(
    text: str, avatar_id: str, engine: str, user: str, uid: str, role: str, out_path: str,
    emotion: str = "neutral", fade_in: bool = False, fade_out: bool = False,
) -> bool:
    """单镜数字人口播：调 digital_human._generate_one 生成人像视频并转竖屏 720x1280。

    返回 True=成功；失败返回 False（由上层回退背景图模式）；
    配额超限（402）向上抛出，由上层整体切换背景图避免逐镜空转。
    emotion（v13.24）：每镜情绪，驱动 TTS 风格 + 表情渲染。
    v13.32 竖屏化升级：纯色 pad 深色底 → 模糊填充背景（split 放大模糊底 + 原画居中
    overlay，无黑边，与插画镜全屏风格统一）；fade_in/fade_out 按镜序控制与
    插画/素材镜对齐（首镜淡入、末镜淡出、中间镜硬切，消除镜间黑场闪烁）。
    """
    try:
        from digital_human import GenerateRequest, _generate_one

        req = GenerateRequest(
            text=text[:5000],
            avatar_id=avatar_id or "business-female",
            resolution="720p",
            engine=engine,
            watermark=False,  # 短剧为平台内容线，不打数字人水印
            emotion=emotion,
        )
        result = _generate_one(req, user, uid, role)
        if result.get("status") != "done" or not result.get("video_url"):
            logger.warning(f"数字人镜头未出片: {result.get('error') or result.get('status')}")
            return False
        src = os.path.join(Path(__file__).parent, result["video_url"].lstrip("/"))
        if not os.path.exists(src):
            logger.warning(f"数字人镜头产物缺失: {src}")
            return False
        # 横屏 1280x720 → 竖屏 720x1280：模糊填充背景（无黑边）+ 可选首尾淡入淡出
        # （v13.32 替代纯色 pad；编码参数与 _scene_video 对齐便于 concat）
        vf = (
            "split=2[bg][fg];"
            "[bg]scale=720:1280:force_original_aspect_ratio=increase,crop=720:1280,"
            "gblur=sigma=25,eq=brightness=-0.10:saturation=0.9[bg];"
            "[fg]scale=720:1280:force_original_aspect_ratio=decrease[fg];"
            "[bg][fg]overlay=(W-w)/2:(H-h)/2,format=yuv420p"
        )
        if fade_in:
            vf += ",fade=t=in:st=0:d=0.25"
        if fade_out:
            dur = _probe_seconds(src)
            vf += f",fade=t=out:st={max(0.25, dur - 0.25):.2f}:d=0.25"
        cmd = [
            FFMPEG_BIN, "-nostdin", "-y", "-i", src,
            "-vf", vf,
            "-c:v", "libx264", "-preset", "fast", "-crf", "20", "-pix_fmt", "yuv420p",
            "-c:a", "aac", "-b:a", "128k", "-r", str(FPS),
            out_path,
        ]
        r = subprocess.run(cmd, capture_output=True, timeout=300)
        if r.returncode != 0 or not os.path.exists(out_path) or os.path.getsize(out_path) < 4096:
            logger.warning("数字人镜头竖屏转换失败: " + r.stderr.decode(errors="replace")[-200:])
            return False
        return True
    except HTTPException as e:
        if e.status_code == 402:
            raise  # 配额超限：上层整体切换背景图模式
        logger.warning(f"数字人镜头失败(HTTP {e.status_code}): {e.detail}")
        return False
    except Exception as e:
        logger.warning(f"数字人镜头失败: {e}")
        return False



def _prepare_drama_context(task_id, drama_config):
    """准备短剧生成上下文。"""
    return {
        "task_id": task_id,
        "config": drama_config,
        "scenes": [],
        "status": "prepared"
    }

def _generate_drama_scene(scene_index, script_data, visual_style):
    """生成单个短剧场景。"""
    return {
        "index": scene_index,
        "script": script_data,
        "style": visual_style,
        "status": "generated"
    }

def _finalize_drama_result(scenes):
    """汇总短剧生成结果。"""
    return {
        "total_scenes": len(scenes),
        "scenes": scenes,
        "status": "completed"
    }


def _drama_generate_simple(drama_params: dict) -> dict:
    """简化版：生成短剧视频。"""
    # 简化的生成逻辑
    return {
        "status": "success",
        "video_url": drama_params.get("output_path", ""),
        "duration": drama_params.get("duration", 0)
    }

def _prepare_drama_params(request_data: dict) -> dict:
    """简化版：准备短剧生成参数。"""
    return {
        "script": request_data.get("script", ""),
        "style": request_data.get("style", ""),
        "output_path": request_data.get("output_path", "")
    }

def _drama_quota_check(uid: str, avatar_mode: bool) -> None:
    """额度检查：经典动画卡模式 worker 内扣费。"""
    if not avatar_mode:
        from common.auth import consume_quota

        quota = consume_quota(uid)
        if not quota.get("allowed"):
            raise HTTPException(
                402,
                "今日短剧生成次数已用完，可在次日 0 点自动恢复",
            )


async def _drama_load_script(theme: str, scenes_override: list, duration_hint: int, payload: dict) -> dict:
    """剧本：自定义分镜优先，否则 LLM 生成（模板注入 + 时长防御）。"""
    script = {"title": payload.get("title") or "未命名短剧", "scenes": scenes_override} if scenes_override else None
    if script is None:
        tpl = None
        tid = payload.get("template_id") or ""
        if tid:
            try:
                from common.template_utils import load_one
                from drama_templates import TEMPLATE_DIR

                tpl = load_one(TEMPLATE_DIR, tid, "题材模板不存在")
            except Exception:  # noqa: BLE001
                logger.warning(f"题材模板加载失败：{tid}")
        script = await _generate_script(theme, duration_hint, tpl)
        if tpl:
            try:
                from drama_templates import record_usage

                record_usage(tid)
            except Exception:  # noqa: BLE001
                pass
    scenes = _enforce_duration(script["scenes"], duration_hint)
    script["scenes"] = scenes
    return script



async def _drama_render_one(
    i: int, sc: dict, text: str, tmpdir: str, avatar_mode: bool, avatar_id: str, dh_engine: str,
    user: str, uid: str, role: str, illust_mode: bool, char_map: dict, char_refs: dict,
    dh_off: bool, fade_in: bool, fade_out: bool, motion: str, title: str, _report, total: int,
    art_style: str = "", last_frame: bytes | None = None, dynamic_on: bool = True,
) -> tuple:
    """dynamic_on: 本场是否启用 i2v 动态锚（render_scenes 按 dynamic_level 算好传入）。"""
    """单镜渲染：数字人 → 素材 → 插画/卡片 三级回退。返回 (clip, audio_path, dh_off)。"""
    clip = os.path.join(tmpdir, f"seg_{i:03d}.mp4")
    audio_path = ""
    ok = False
    if avatar_mode and not dh_off:
        try:
            ok = await asyncio.to_thread(
                _dh_scene_video, text, avatar_id, dh_engine, user, uid, role, clip,
                sc.get("emotion", "neutral"), fade_in, fade_out,
            )
        except HTTPException as e:
            if e.status_code == 402:
                dh_off = True
                ok = False
                _report(15 + int(50 * i / max(total, 1)), "数字人额度不足，切换素材模式…")
            else:
                raise
    if not ok:
        audio = await asyncio.to_thread(_tts_scene, text, sc.get("emotion", "neutral"))
        audio_path = os.path.join(tmpdir, f"seg_{i:03d}.mp3")
        with open(audio_path, "wb") as f:
            f.write(audio)
        dur = max(_probe_seconds(audio_path), float(sc.get("sec") or 5))
        if not illust_mode:
            scene_chars = [c for c in (sc.get("chars") or []) if c in char_map]
            lead = char_map.get(scene_chars[0]) if scene_chars else None
            search_q = _anchor_search(lead, sc.get("search", ""))
            ok = await asyncio.to_thread(_material_scene_video, search_q, audio_path, clip, dur, fade_in, fade_out)
    if not ok:
        img_path = os.path.join(tmpdir, f"seg_{i:03d}.jpg")
        shot = sc.get("shot", "")
        data = None
        if shot:
            scene_chars = [c for c in (sc.get("chars") or []) if c in char_map]
            _sc = [c for c in scene_chars if char_map[c].get("anchor")]
            if len(_sc) >= 2:
                # 多角色同框：按出场顺序标注位置（左/右），模型对位参考图
                _pos = ["左边", "右边", "中间"][:len(_sc)]
                anchors = "；".join(
                    f"{_pos[i]}的是{char_map[c]['name']}（{char_map[c]['anchor']}）" for i, c in enumerate(_sc)
                )
            else:
                anchors = "、".join(char_map[c]["anchor"] for c in _sc)
            refs = [char_refs[c] for c in scene_chars if c in char_refs]
            # v1.0.46 双参考：角色立绘（同脸）+ 上一镜画面（镜间连续性：场景/光线/服装延续）
            _refs2 = list(refs)
            if last_frame and _refs2:
                _refs2.append(last_frame)
            data = await asyncio.to_thread(
                _generate_scene_image, shot, anchors, _refs2, uid, art_style,
                sc.get("dialogue") or "", sc.get("shot_size") or "",
                resolve_api_key(), resolve_api_base(),
            )
        if data:
            with open(img_path, "wb") as f:
                f.write(data)
            # v1.0.41 场次拆分子镜头：一场戏按目标镜头时长切成多个子镜头
            # （同一主图 + 不同机位/取景窗口 → 视觉连续感，对标红果漫剧节奏）
            try:
                shot_dur = float(sc.get("sec") or dur)
                # 目标子镜头 3.2s（短视频节奏），一场 15-40s → 5-12 个子镜
                sub_target = 3.2
                n_sub = max(1, min(12, int(round(shot_dur / sub_target))))
                if n_sub <= 1:
                    await asyncio.to_thread(_scene_video, img_path, audio_path, clip, dur, motion, fade_in, fade_out)
                else:
                    sub_audios = _split_audio_segments(audio_path, n_sub)
                    if len(sub_audios) < 2:
                        await asyncio.to_thread(_scene_video, img_path, audio_path, clip, dur, motion, fade_in, fade_out)
                    else:
                        sub_clips = []
                        _shot_seq = _shot_sequence(sc.get("emotion") or "", len(sub_audios), i)
                        for si, sa in enumerate(sub_audios):
                            sclip = os.path.join(tmpdir, f"shot_{i:03d}_{si:02d}.mp4")
                            # 子镜时长撑满场次目标（红果漫剧：画面节奏独立于配音，配音短则画面留白/慢镜）
                            s_dur = max(shot_dur / n_sub, _probe_seconds(sa), 2.0)
                            # 机位递进（场次内景别渐进：交代→进入→聚焦→收尾）
                            shot_type = _shot_seq[si % len(_shot_seq)]
                            # 取景窗口：全图或局部（特写窗口偏上中、横移窗口偏左/右）
                            if shot_type == "close_zoom":
                                win = (0.15, 0.15, 0.7, 0.7)  # 中心偏上：面部
                            elif shot_type == "pan_left":
                                win = (0.0, 0.0, 0.8, 1.0)   # 左侧起始，右扫
                            elif shot_type == "pan_right":
                                win = (0.2, 0.0, 0.8, 1.0)   # 右侧起始，左扫
                            elif shot_type == "tilt_up":
                                win = (0.1, 0.2, 0.8, 0.8)   # 偏下起始，上移
                            elif shot_type == "tilt_down":
                                win = (0.1, 0.0, 0.8, 0.8)   # 偏上起始，下移
                            else:
                                win = (0.0, 0.0, 1.0, 1.0)   # 全图推拉
                            # 每个子镜带 0.1s 微淡（平滑镜头切换，防取景窗口跳变生硬）；
                            # 场首子镜用整场 fade_in、场末子镜用 fade_out
                            await asyncio.to_thread(
                                _scene_video, img_path, sa, sclip, s_dur,
                                shot_type, True, True, win,
                            )
                            if os.path.exists(sclip) and os.path.getsize(sclip) > 4096:
                                sub_clips.append(sclip)
                        # v1.0.45 动态锚镜头：本场有台词/人物时，用主图生成 1 个 i2v
                        # 动态片段（人物真动/口型），替换第一个静态子镜 → 画面"活"起来
                        if dynamic_on and sc.get("dialogue") and len(sub_clips) >= 2:
                            _report(15 + int(50 * i / max(total, 1)), f"第 {i + 1}/{total} 镜：动态镜头生成中…")
                            dyn_clip = os.path.join(tmpdir, f"dyn_{i:03d}.mp4")
                            _motion_p = _i2v_motion_prompt(sc.get("emotion") or "", sc.get("dialogue") or "")
                            _ctx_key = resolve_api_key()
                            _ctx_base = resolve_api_base()
                            dyn_ok = await asyncio.to_thread(
                                _i2v_scene_clip, img_path, _motion_p, dyn_clip, uid, 240,
                                _ctx_key, _ctx_base,
                            )
                            if not dyn_ok:
                                logger.warning(f"i2v 动态锚失败（第 {i+1} 镜），回退静态")
                            if dyn_ok and os.path.exists(dyn_clip) and os.path.getsize(dyn_clip) > 10000:
                                # 动态锚替换第一个静态子镜，并把该子镜配音合入动态画面
                                _dyn_audio = sub_audios[0] if len(sub_audios) >= 1 else None
                                _dyn_final = os.path.join(tmpdir, f"dyn_final_{i:03d}.mp4")
                                _mux_ok = False
                                if _dyn_audio and os.path.exists(_dyn_audio):
                                    _mux_ok = await asyncio.to_thread(
                                        _mix_dyn_audio, dyn_clip, _dyn_audio, _dyn_final,
                                    )
                                if _mux_ok and os.path.exists(_dyn_final):
                                    sub_clips[0] = _dyn_final
                                else:
                                    sub_clips[0] = dyn_clip
                        if len(sub_clips) >= 2:
                            _concat_sub_shots(sub_clips, clip)
                        elif sub_clips:
                            import shutil as _sh
                            _sh.copyfile(sub_clips[0], clip)
                        else:
                            await asyncio.to_thread(_scene_video, img_path, audio_path, clip, dur, motion, fade_in, fade_out)
                ok = True
            except Exception as e:
                logger.warning(f"子镜头拆分失败，回退单镜: {e}")
                await asyncio.to_thread(_scene_video, img_path, audio_path, clip, dur, motion, fade_in, fade_out)
                ok = True
        else:
            ok = await asyncio.to_thread(_make_scene_card, text, i, total, title, img_path, uid=uid)
            if ok:
                await asyncio.to_thread(_scene_video, img_path, audio_path, clip, dur, "still", fade_in, fade_out)
    return (clip if ok else None), audio_path, dh_off

async def _drama_render_scenes(scenes: list, payload: dict, user: str, uid: str, role: str, tmpdir: str, _report) -> tuple:
    """逐镜配音 + 画面（三级回退：数字人 → 素材 → 插画/卡片）。返回 (clip_paths, srt_durations, voice_durations)。"""
    avatar_mode = bool(payload.get("avatar_mode"))
    avatar_id = (payload.get("avatar_id") or "business-female").strip()
    dh_engine = (payload.get("dh_engine") or "2d").strip()
    illust_mode = bool(payload.get("illust_mode"))
    art_style = (payload.get("art_style") or DEFAULT_ART_STYLE).strip().lower()
    if art_style not in ART_STYLES:
        art_style = DEFAULT_ART_STYLE
    _dyn_lv = (payload.get("dynamic_level") or "auto").strip().lower()
    characters = payload.get("characters") or []
    clip_paths, srt_durations, voice_durations = [], [], []
    total = len(scenes)
    dh_off = False
    char_map = {c.get("id"): c for c in characters if c.get("id")}
    char_refs: dict[str, bytes] = {}
    # 漫剧模式：渲染前为每个出场角色预热定妆立绘（全剧同脸同装核心）。
    # 立绘作为每镜图生图参考锚定；角色圣经已存立绘时直接复用（跨集同脸）。
    if illust_mode and char_map:
        _report(12, "角色定妆中…（保证全剧角色一致）")
        for _cid, _char in char_map.items():
            _portrait = _load_char_portrait(_cid, _char, uid, art_style, resolve_api_key(), resolve_api_base())
            if _portrait:
                char_refs[_cid] = _portrait
                # 单角色立绘缓存到本地，跨集复用
                _save_char_portrait(_cid, _char, _portrait, art_style)
    scene_bounds: list[int] = []  # 每个场次第一个 clip 的下标（供场间转场）
    _last_scene_frame: bytes | None = None  # 上一镜成功画面（双参考：镜间连续性）
    for i, sc in enumerate(scenes):
        _report(15 + int(50 * i / max(total, 1)), f"第 {i + 1}/{total} 镜：配音 + 画面…")
        text = " ".join(x for x in (sc.get("narrator"), sc.get("dialogue")) if x)
        if not text:
            continue
        clip = os.path.join(tmpdir, f"seg_{i:03d}.mp4")
        scene_chars = [c for c in (sc.get("chars") or []) if c in char_map]
        fade_in, fade_out = i == 0, i == total - 1
        motion = _motion_for(sc, i)
        _last_frame = _last_scene_frame
        _dyn_on = _dyn_lv == "on" or (_dyn_lv == "auto" and i % 2 == 0)
        result = await _drama_render_one(
            i, sc, text, tmpdir, avatar_mode, avatar_id, dh_engine, user, uid, role,
            illust_mode, char_map, char_refs, dh_off, fade_in, fade_out, motion,
            payload.get("title") or "未命名短剧", _report, total, art_style, _last_frame, _dyn_on,
        )
        clip, audio_path, dh_off = result
        if clip:
            # 双参考：本场插画画面作为下一镜的镜间连续性参考
            if illust_mode:
                _imgp = os.path.join(tmpdir, f"seg_{i:03d}.jpg")
                if os.path.exists(_imgp) and os.path.getsize(_imgp) > 10000:
                    try:
                        _last_scene_frame = open(_imgp, "rb").read()
                    except Exception:
                        pass
            if not clip_paths:
                scene_bounds.append(0)
            clip_paths.append(clip)
            srt_durations.append(_probe_seconds(clip))
            vd = _probe_seconds(audio_path) if audio_path else _probe_seconds(clip)
            voice_durations.append(vd)
        elif clip_paths:
            # 本场无画面（异常跳过）→ 下一场从当前 clip 数开始（保持边界正确）
            pass
    return clip_paths, srt_durations, voice_durations, scene_bounds


async def _drama_generate_worker(payload: dict, progress: Callable | None = None) -> dict:  # noqa: C901
    """短剧生成执行体：剧本 → 配音 → 镜头图 → 合成 → 字幕。"""

    def _report(pct: float, stage: str) -> None:
        _notify_progress(progress, pct, stage)

    theme = (payload.get("theme") or "").strip()
    scenes_override = payload.get("scenes") or []
    if not theme and not scenes_override:
        raise HTTPException(400, "请输入短剧主题")
    title = (payload.get("title") or "").strip() or "未命名短剧"
    user = payload.get("user") or ""
    uid = payload.get("uid") or ""
    role = payload.get("role") or ""
    avatar_mode = bool(payload.get("avatar_mode"))

    _drama_quota_check(uid, avatar_mode)

    tmpdir = tempfile.mkdtemp(prefix="drama_")
    try:
        # 1. 剧本
        _report(5, "剧本创作中…")
        duration_hint = max(20, min(1800, int(payload.get("duration") or 45)))
        script = await _drama_load_script(theme, scenes_override, duration_hint, payload)
        scenes = script["scenes"]

        # 2. 逐镜配音 + 画面（三级回退）
        _report(15, "分镜配音与画面生成中…")
        clip_paths, srt_durations, voice_durations, scene_bounds = await _drama_render_scenes(
            scenes, payload, user, uid, role, tmpdir, _report
        )
        if not clip_paths:
            raise HTTPException(502, "所有分镜合成失败，请重试")

        # 2.5 片头标题卡（漫剧标准开场：剧名+集数，3.5s，BGM 垫底）
        intro_clip = ""
        if payload.get("illust_mode"):
            _report(70, "片头制作中…")
            _intro_style = (payload.get("art_style") or DEFAULT_ART_STYLE).strip().lower()
            if _intro_style not in ART_STYLES:
                _intro_style = DEFAULT_ART_STYLE
            intro_path = await asyncio.to_thread(
                _make_intro_card, title, _intro_style,
                f"第 {payload.get('episode') or 1} 集" if payload.get('episode') else "",
            )
            if intro_path:
                try:
                    intro_clip = os.path.join(tmpdir, "intro.mp4")
                    _bgm = _pick_bgm()
                    if _bgm:
                        # 片头：标题卡 + BGM 前奏（2s 淡入）
                        _report(71, "片头配乐中…")
                        await asyncio.to_thread(
                            _make_intro_video, intro_path, _bgm, intro_clip, 3.5,
                        )
                    else:
                        await asyncio.to_thread(
                            _scene_video, intro_path, None, intro_clip, 3.5,
                            "slow_push", True, True, (0.0, 0.0, 1.0, 1.0),
                        )
                except Exception as e:
                    logger.warning(f"片头生成失败: {e}")
                    intro_clip = ""

        # 2.6 片尾卡（红果漫剧结尾：剧名 + 下集预告提示）
        outro_clip = ""
        if illust_mode:
            try:
                outro_path = await asyncio.to_thread(
                    _make_intro_card, title, _intro_style,
                    "本集完 · 下集更精彩",
                )
                if outro_path:
                    outro_clip = os.path.join(tmpdir, "outro.mp4")
                    await asyncio.to_thread(
                        _scene_video, outro_path, None, outro_clip, 3.0,
                        "slow_push", True, True, (0.0, 0.0, 1.0, 1.0),
                    )
            except Exception as e:
                logger.warning(f"片尾生成失败: {e}")
                outro_clip = ""

        # 3. 拼接（片头 + 正片 + 片尾，场次间交叉淡化转场）+ 字幕
        _report(72, "片段拼接中…")
        raw_video = os.path.join(tmpdir, "merged.mp4")
        if intro_clip and os.path.exists(intro_clip):
            clip_paths.insert(0, intro_clip)
            scene_bounds = [b + 1 for b in scene_bounds]
            scene_bounds.insert(0, 0)
        if outro_clip and os.path.exists(outro_clip):
            clip_paths.append(outro_clip)
        await asyncio.to_thread(_concat_videos, clip_paths, raw_video, scene_bounds)
        stem = f"drama_{int(time.time() * 1000)}"
        srt_path = os.path.join(tmpdir, f"{stem}.srt")
        _make_srt(scenes[: len(srt_durations)], srt_durations, voice_durations, srt_path)
        final_name = f"{stem}.mp4"
        final_path = DRAMA_DIR / final_name
        _report(88, "字幕合成中…")
        # 竖屏字幕安全区：底部 5% 边距，字幕不贴底不挡脸（红果短剧规范）
        await asyncio.to_thread(_burn_subtitles, raw_video, srt_path, str(final_path), _pick_bgm(), margin_v=36)
        shutil.copyfile(srt_path, DRAMA_DIR / f"{stem}.srt")
        cover_path = DRAMA_DIR / f"{stem}.jpg"
        preview_path = DRAMA_DIR / f"{stem}_preview.mp4"
        await asyncio.to_thread(_extract_cover, clip_paths[0], str(cover_path))
        await asyncio.to_thread(_make_preview, clip_paths[0], str(preview_path))

        # 成片 QC 质量门（不阻塞，仅标记问题）
        _report(94, "质量检查中…")
        qc = await asyncio.to_thread(_qc_check, final_path, srt_path)

        duration = _probe_seconds(str(final_path))
        cover_url = f"/api/drama/covers/{stem}.jpg"
        art_id = save_artifact(
            art_type="video",
            author="short_drama",
            media_url=f"/api/drama/videos/{final_name}",
            content=theme,
            metadata={
                "title": script["title"],
                "theme": theme,
                "scenes": len(srt_durations),
                "duration": duration,
                "engine": "dh" if avatar_mode else "local",
                "avatar_mode": avatar_mode,
                "cover": cover_url,
            },
            duration=duration,
            thumbnail=cover_url,
        )
        _report(100, "短剧已生成")
        return {
            "id": final_name,
            "artifact_id": art_id,
            "url": f"/api/drama/videos/{final_name}",
            "srt_url": f"/api/drama/srt/{stem}.srt",
            "cover_url": cover_url,
            "title": script["title"],
            "theme": theme,
            "scenes": len(srt_durations),
            "duration": duration,
            "avatar_mode": avatar_mode,
            "engine": "dh" if avatar_mode else "local",
            "qc": qc,
        }
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

def _drama_handler(task_id: str, payload: dict, update: Callable, ctx: dict) -> dict:
    """异步任务 handler（register_handler 约定签名）。"""

    async def run() -> dict:
        return await _drama_generate_worker(payload, update)

    return asyncio.run(run())


register_handler("drama_generate", _drama_handler, user_limit=1, pool="long", max_attempts=2)


@router.get("/config")
async def drama_config(current_user: dict = require_auth()):
    """素材源状态（v13.25）：前端据此提示 Pexels key / 本地素材 / BGM 就绪情况。"""
    local_count = 0
    if MATERIALS_DIR.exists():
        local_count = sum(
            1 for p in MATERIALS_DIR.rglob("*")
            if p.is_file() and p.suffix.lower() in _VIDEO_EXTS + _IMAGE_EXTS
        )
    music_count = sum(1 for p in MUSIC_DIR.glob("*") if p.is_file() and p.suffix.lower() in _MUSIC_EXTS)
    return {
        "pexels_configured": bool(_resolve_pexels_key()),
        "local_materials": local_count,
        "music_tracks": music_count,
    }


# ─── v15 分镜表导出 + 素材清单（纯函数，供端点与单测复用）───
_EMOTION_CN = {
    "neutral": "自然",
    "happy": "欢快",
    "sad": "悲伤",
    "angry": "激昂",
    "gentle": "温柔",
    "serious": "严肃",
}

_SHOT_SHEET_COLS = ["镜号", "时长(秒)", "情绪", "出场角色", "画面描述(shot)", "素材关键词(search)", "旁白(narrator)", "台词(dialogue)"]


def build_shot_sheet(scenes: list[dict], title: str = "", characters: list[dict] | None = None) -> bytes:
    """分镜表 xlsx（openpyxl）：标题行（可选）+ 表头 + 每镜一行。纯函数，返回文件字节。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    char_map = {str(c.get("id")): str(c.get("name") or c.get("id")) for c in (characters or []) if c.get("id")}
    wb = Workbook()
    ws = wb.active
    ws.title = "分镜表"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7C3AED")
    if title:
        ws.append([f"《{title}》分镜表"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_SHOT_SHEET_COLS))
        ws.cell(1, 1).font = Font(bold=True, size=13)
        ws.append([])  # 空行分隔标题与表头
    ws.append(list(_SHOT_SHEET_COLS))
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, s in enumerate(scenes or [], 1):
        chars = "、".join(char_map.get(cid, cid) for cid in (s.get("chars") or []))
        ws.append(
            [
                i,
                int(s.get("sec") or 5),
                _EMOTION_CN.get(str(s.get("emotion") or "").lower(), s.get("emotion") or "自然"),
                chars,
                s.get("shot") or "",
                s.get("search") or "",
                s.get("narrator") or "",
                s.get("dialogue") or "",
            ]
        )
    # 列宽自适应（按内容上限 50 字符，避免超宽）
    for idx, col in enumerate(_SHOT_SHEET_COLS, 1):
        letter = get_column_letter(idx)
        max_len = max((len(col) * 2 + 4, 10))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max_len, 50)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_material_manifest(scenes: list[dict]) -> dict:
    """批量生成素材清单：每镜素材需求（关键词/时长/情绪/文案）+ 汇总统计 + md。纯函数。"""
    items = []
    for i, s in enumerate(scenes or [], 1):
        keyword = (s.get("search") or "").strip()
        shot = (s.get("shot") or "").strip()
        text = " ".join(x for x in (s.get("narrator"), s.get("dialogue")) if x)
        items.append(
            {
                "no": i,
                "keyword": keyword or shot[:30] or f"scene_{i}",
                "sec": int(s.get("sec") or 5),
                "emotion": _EMOTION_CN.get(str(s.get("emotion") or "").lower(), "自然"),
                "text": text,
                "text_len": len(text),
            }
        )
    keywords: list[str] = []
    for it in items:
        for kw in re.split(r"[,，;；/\\|\s]+", it["keyword"]):
            kw = kw.strip()
            if kw and kw not in keywords:
                keywords.append(kw)
    summary = {
        "total_scenes": len(items),
        "total_sec": sum(it["sec"] for it in items),
        "total_text_chars": sum(it["text_len"] for it in items),
        "keywords": keywords,
    }
    lines = [
        "# 短剧素材清单",
        "",
        f"共 {len(items)} 镜，总时长约 {summary['total_sec']} 秒，台词/旁白共 {summary['total_text_chars']} 字",
        "",
        "| 镜号 | 素材关键词 | 建议时长 | 情绪 | 文案字数 |",
        "| --- | --- | --- | --- | --- |",
    ]
    lines += [f"| {it['no']} | {it['keyword']} | {it['sec']}s | {it['emotion']} | {it['text_len']} |" for it in items]
    lines += ["", "## 关键词汇总", ""]
    lines += [f"- {kw}" for kw in keywords] or ["- （无）"]
    lines += [
        "",
        "## 使用说明",
        "- 素材模式：将关键词命名的素材（*关键词*.mp4/jpg）放入 backend/drama_factory/materials/ 目录，生成时自动匹配",
        "- Pexels Key 已配置时优先实时搜索真实视频素材，本地素材作为兜底",
        "- 配音不足时画面按 sec 循环补足，建议素材时长 ≥ 对应镜长（8-40 秒为佳）",
    ]
    return {"items": items, "summary": summary, "manifest_md": "\n".join(lines)}


@router.post("/export-shots")
async def export_shot_sheet(
    title: str = Form(""),
    scenes_json: str = Form(""),
    characters_json: str = Form(""),
    current_user: dict = require_auth(),
):
    """导出分镜表 Excel（xlsx）：每镜一行（时长/情绪/角色/画面/关键词/台词）。"""
    try:
        scenes = json.loads(scenes_json or "[]")
        if not isinstance(scenes, list) or not scenes:
            raise ValueError("分镜为空")
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(400, "分镜 JSON 格式错误，请检查 scenes_json 是否符合 [{shot,narrator,dialogue,sec}] 结构") from e
    characters = []
    if characters_json:
        try:
            characters = json.loads(characters_json)
        except json.JSONDecodeError as e:
            raise HTTPException(400, "服务异常，请稍后重试") from e
    data = build_shot_sheet(scenes, title.strip(), characters)
    from urllib.parse import quote

    filename = f"{title.strip() or '短剧'}-分镜表.xlsx"
    try:
        filename.encode("latin-1")
        ascii_name = filename
    except UnicodeEncodeError:
        ascii_name = "drama-shots.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'},
    )


@router.post("/material-manifest")
async def material_manifest(
    scenes_json: str = Form(""),
    current_user: dict = require_auth(),
):
    """批量生成素材清单：每镜素材需求（关键词/时长/情绪）+ 汇总统计 + 清单 md。"""
    try:
        scenes = json.loads(scenes_json or "[]")
        if not isinstance(scenes, list) or not scenes:
            raise ValueError("分镜为空")
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(400, "分镜 JSON 格式错误，请检查 scenes_json 是否符合 [{shot,narrator,dialogue,sec}] 结构") from e
    return build_material_manifest(scenes)


@router.post("/script")
async def generate_script(
    theme: str = Form(""),
    duration: int = Form(45),
    template_id: str = Form("", description="题材模板 ID（drama-templates，如 dt_ceo）"),
    current_user: dict = require_auth(),
):
    """AI 写剧本（v13.29 + v22 题材模板）：主题 + 目标时长 + 可选题材模板 → 剧本 JSON。

    返回的 scenes 可直接作为 /generate 的 scenes_json 提交——前端剧本工作台
    编辑后确认生成，保证"所见即所得"（返回即最终成片剧本，已过时长防御）。
    """
    theme = theme.strip()
    if not theme:
        raise HTTPException(400, "请输入短剧主题")
    duration_hint = max(20, min(1800, int(duration) or 45))
    tpl = None
    if template_id:
        try:
            from common.template_utils import load_one
            from drama_templates import TEMPLATE_DIR

            tpl = load_one(TEMPLATE_DIR, template_id, "题材模板不存在")
        except Exception:  # noqa: BLE001
            raise HTTPException(404, "题材模板不存在") from None
    script = await _generate_script(theme, duration_hint, tpl)
    return {
        "title": script["title"],
        "scenes": script["scenes"],
        "characters": script.get("characters") or [],  # v13.30 角色表（角色一致性）
        "template_id": template_id or "",
    }


@router.post("/generate")
async def generate_drama(
    theme: str = Form(""),
    title: str = Form(""),
    duration: int = Form(45),
    scenes_json: str = Form(""),
    characters_json: str = Form("", description="角色表 JSON（[{id,name,gender,age,appearance,outfit,search}]）"),
    template_id: str = Form("", description="题材模板 ID（drama-templates，如 dt_ceo）"),
    illust_mode: bool = Form(False, description="true=AI 插画模式（AGNES 文生图/图生图，角色一致性）"),
    art_style: str = Form("", description="画风预设：guoman国漫/hanman韩漫/3d/realistic写实（漫剧模式）"),
    dynamic_level: str = Form("auto", description="动态镜头级别：auto自动/on开启/off关闭（i2v 人物动态，慢且耗配额）"),
    avatar_mode: bool = Form(False, description="true=数字人播报模式（每镜生成人像口播视频）"),
    avatar_id: str = Form("business-female", description="数字人形象ID（avatar_mode 时生效）"),
    dh_engine: str = Form("2d", description="数字人引擎：2d/live_portrait（sadtalker 耗时过长不适用）"),
    sync: bool = Query(False, description="true=同步执行（脚本/测试用）；默认异步任务"),
    current_user: dict = require_auth(),
):
    """生成短剧（默认异步任务，立即返回 task_id）。

    - theme: 主题（LLM 自动生成剧本分镜）
    - scenes_json: 可选自定义分镜 JSON（[{shot,narrator,dialogue,sec}]）
    - characters_json: 可选角色表 JSON（v13.30，角色一致性锚定）
    - illust_mode: AI 插画模式（每镜文生图/图生图，角色参考图锚定同人）
    - avatar_mode: 数字人播报模式（每镜生成人像口播视频，失败自动回退背景图）
    - 本地管线：CosyVoice 配音 + 画面 + 字幕，无外部视频 API 依赖
    """
    scenes = []
    if scenes_json:
        try:
            scenes = json.loads(scenes_json)
        except json.JSONDecodeError as e:
            raise HTTPException(400, "服务异常，请稍后重试") from e
    characters = []
    if characters_json:
        try:
            characters = json.loads(characters_json)
            if not isinstance(characters, list):
                raise ValueError("characters 必须是数组")
        except (json.JSONDecodeError, ValueError) as e:
            raise HTTPException(400, "服务异常，请稍后重试") from e
    if not theme.strip() and not scenes:
        raise HTTPException(400, "请输入短剧主题或提供自定义分镜")
    if avatar_mode and dh_engine not in ("2d", "live_portrait"):
        raise HTTPException(400, "短剧数字人引擎仅支持 2d / live_portrait（sadtalker 耗时过长不适用）")
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    uid = current_user.get("user_id", "") if isinstance(current_user, dict) else ""
    role = current_user.get("role", "") if isinstance(current_user, dict) else ""
    payload = {
        "theme": theme,
        "title": title,
        "duration": duration,
        "scenes": scenes,
        "characters": characters,
        "template_id": template_id,
        "illust_mode": illust_mode,
        "art_style": art_style,
        "dynamic_level": dynamic_level,
        "avatar_mode": avatar_mode,
        "avatar_id": avatar_id,
        "dh_engine": dh_engine,
        "user": user,
        "uid": uid,
        "role": role,
    }
    if sync:
        return await _drama_generate_worker(payload)
    task = create_task("drama_generate", payload, username=user, user_id=uid, role=role)
    return {
        "task_id": task["id"],
        "status": "pending",
        "message": "短剧创作任务已提交（剧本 + 配音 + 视频合成，约 2-15 分钟，随时长增加）",
        "task": task,
    }


@router.get("/videos/{filename}")
async def get_video(filename: str):
    """下载/播放短剧视频。"""
    path = DRAMA_DIR / filename
    if not path.exists():
        raise HTTPException(404, "视频不存在")
    return FileResponse(path, media_type="video/mp4")


@router.get("/srt/{filename}")
async def get_srt(filename: str):
    """下载字幕文件。"""
    path = DRAMA_DIR / filename
    if not path.exists():
        raise HTTPException(404, "字幕不存在")
    return FileResponse(path, media_type="application/x-subrip")


@router.get("/covers/{filename}")
async def get_cover(filename: str):
    """获取封面图。"""
    path = DRAMA_DIR / filename
    if not path.exists():
        raise HTTPException(404, "封面不存在")
    return FileResponse(path, media_type="image/jpeg")


@router.get("/list")
async def list_dramas(current_user: dict = require_auth()):
    """短剧作品列表（按时间倒序）。v13.25：标题从 artifacts 表 metadata 补全。"""
    # 生成时已登记 artifacts（author=short_drama, type=video），metadata 含 title/theme
    titles: dict[str, str] = {}
    try:
        from common.db import get_db_context

        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT media_url, metadata FROM artifacts WHERE type='video' AND author='short_drama'"
            ).fetchall()
        for media_url, metadata in rows:
            try:
                md = json.loads(metadata or "{}")
                if md.get("title"):
                    titles[Path(media_url or "").name] = md["title"]
            except (TypeError, json.JSONDecodeError):
                continue
    except Exception as e:
        logger.debug(f"读取短剧标题失败: {e}")
    items = []
    if DRAMA_DIR.exists():
        # v13.28 过滤首镜预览片段（_preview.mp4），避免被当作独立作品展示
        files = sorted(
            (f for f in DRAMA_DIR.glob("drama_*.mp4") if not f.name.endswith("_preview.mp4")),
            reverse=True,
        )
        for f in files:
            items.append(
                {
                    "id": f.name,
                    "title": titles.get(f.name, ""),
                    "url": f"/api/drama/videos/{f.name}",
                    "srt_url": f"/api/drama/srt/{f.stem}.srt",
                    "cover_url": f"/api/drama/covers/{f.stem}.jpg",
                    "preview_url": f"/api/drama/videos/{f.stem}_preview.mp4",
                    "duration": _probe_seconds(str(f)),
                    "created_at": time.strftime("%Y-%m-%d %H:%M", time.localtime(f.stat().st_mtime)),
                }
            )




# ══════════════════════════════════════════════════════════════
# 红果短剧升级：① 小说原文转剧本 ② 角色圣经（跨集一致性）③ 系列连载
# ══════════════════════════════════════════════════════════════

_NOVEL_SYSTEM = """你是资深短剧编剧。把用户提供的小说/故事原文改编成节奏紧凑的竖屏短剧脚本。

要求：
1. 忠实原著：角色名、关键情节、人物关系必须沿用原文；删繁就简，聚焦 1-4 个核心角色和最有冲突感的主线
2. 输出严格的 JSON（不要 markdown 代码块，不要多余文字）
3. 结构：{"title": "剧名", "episode": 1, "characters": [{"id": "lin", "name": "林小满", "gender": "女", "age": "24岁", "appearance": "黑色长发齐刘海，圆脸大眼睛", "outfit": "白色连衣裙配红色围巾", "search": "young chinese woman black hair"}], "scenes": [{"id": 1, "chars": ["lin"], "shot": "镜头画面描述", "search": "英文素材关键词", "narrator": "旁白", "dialogue": "角色台词", "emotion": "情绪", "sec": 25}]}
4. 角色一致性（最重要）：先定义 1-4 个主要角色 characters，每个角色性别/年龄/发型/发色/服装全剧固定；每镜 chars 列出场角色（1-2 个为宜）；每个角色首次出场安排单人镜定妆，后续沿用；shot 必须点名角色并沿用其外貌服装；search 以该镜主角英文特征词开头
5. 场次与目标时长匹配：场次数 ≈ 目标秒数 ÷ 每场 25-30 秒（如 45 秒 → 2-3 场；300 秒 → 10-12 场；最多 28 场）；每场 sec 15-40 秒
6. 每镜必须标注景别 shot_size（特写/近景/中景/全景/远景），按剧情情绪选：情绪激烈用特写近景，交代环境用全景远景，对话用中景；镜头画面 shot 必须与景别一致
7. shot 必须是"画面能拍到的东西"；search 给 2-4 个对应英文关键词（如 night city rain neon），禁止抽象词
8. 每场旁白+台词约 60-100 字；台词必须提到本镜画面里的具体元素，与 shot 强呼应
9. 每镜标注情绪 emotion：happy/sad/angry/gentle/serious/neutral
10. 剧情有起承转合，结尾留悬念钩子（为下一集铺垫）"""


def _parse_characters(data: dict) -> list[dict]:
    """解析剧本 JSON 里的角色表（兼容无角色表）。"""
    chars = data.get("characters") or []
    out = []
    if isinstance(chars, list):
        for c in chars:
            if not isinstance(c, dict) or not c.get("name"):
                continue
            out.append(
                {
                    "id": re.sub(r"[^a-z0-9]", "", str(c.get("id") or "").strip().lower()) or f"c{len(out) + 1}",
                    "name": str(c.get("name"))[:30],
                    "gender": str(c.get("gender") or "女")[:10],
                    "age": str(c.get("age") or "")[:10],
                    "appearance": str(c.get("appearance") or "")[:100],
                    "outfit": str(c.get("outfit") or "")[:100],
                    "search": str(c.get("search") or "")[:80],
                }
            )
    return out




def _repair_json_quotes(candidate: str) -> str:
    """修复 LLM 短剧 JSON 中台词/画面里的裸 ASCII 引号（如 写着"清欢"。）。

    中文内容里混入英文引号会破坏 JSON 结构，此处将其替换为中文引号「」。
    只在「已确定是值内部」时替换：简单策略是替换 { 和 } 之外的成对 ASCII 引号较复杂，
    这里采用安全做法：把「中文语境下成对的 ASCII 引号」替换为中文引号。
    """
    # 成对 ASCII 引号 → 中文引号（只处理中文字符夹着的引号对）
    import re as _re

    # 模式：非 ASCII 引号开头的引号对（"xx"）且两侧是中文/标点 → 中文引号
    out = []
    i = 0
    n = len(candidate)
    while i < n:
        ch = candidate[i]
        if ch == '"':
            # 向前找是否是「中文内容里的引号」：前一个字符是中文或中缀标点，且找到配对引号后跟中文
            prev_is_cjk = i > 0 and (ord(candidate[i-1]) > 0x2E80 or candidate[i-1] in "，。！？、；：）】」…—")
            if prev_is_cjk:
                j = candidate.find('"', i + 1)
                if j != -1:
                    nxt = candidate[j+1] if j + 1 < n else ""
                    next_is_cjk = nxt and (ord(nxt) > 0x2E80 or nxt in "，。！？、；：）】」…—")
                    if next_is_cjk:
                        out.append("\u201c")
                        out.append(candidate[i+1:j].replace('\\"', '"'))
                        out.append("\u201d")
                        i = j + 1
                        continue
        out.append(ch)
        i += 1
    return "".join(out)


def _drama_parse_script(raw: str) -> dict:
    """解析 LLM 剧本 JSON（剥 markdown 代码块/前后噪音/尾随逗号）。"""
    text = raw.strip()
    m = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text, re.S)
    if m:
        text = m.group(1)
    # 兼容：直接输出分镜数组 [{...}]（分块生成时 LLM 可能省略外层对象）
    data = None
    start, end = text.find("{"), text.rfind("}")
    arr_s, arr_e = text.find("["), text.rfind("]")
    candidates = []
    if start >= 0 and end > start:
        candidates.append(text[start : end + 1])
    if arr_s >= 0 and arr_e > arr_s:
        candidates.append(text[arr_s : arr_e + 1])
    for candidate in candidates:
        try:
            data = json.loads(candidate)
            break
        except json.JSONDecodeError:
            try:
                cleaned = re.sub(r",\s*([\]}])", r"\1", candidate)
                data = json.loads(cleaned)
                break
            except json.JSONDecodeError:
                # 中文引号修复后重试（LLM 台词里混入 ASCII 引号）
                try:
                    data = json.loads(_repair_json_quotes(candidate))
                    break
                except json.JSONDecodeError:
                    continue
    if data is None:
        raise ValueError("剧本输出不是 JSON")
    if isinstance(data, list):
        # 直接是分镜数组 → 包一层
        scenes = data
        data = {"scenes": scenes, "characters": []}
    scenes = data.get("scenes") or []
    if not scenes:
        raise ValueError("剧本没有分镜")
    for s in scenes:
        if not s.get("shot") and not s.get("narrator") and not s.get("dialogue"):
            raise ValueError(f"分镜 {s.get('id')} 内容为空")
        s["sec"] = max(2, min(45, int(s.get("sec") or 5)))
        emo = str(s.get("emotion") or "neutral").strip().lower()
        if emo not in ("neutral", "happy", "sad", "angry", "gentle", "serious"):
            emo = {"欢快": "happy", "开心": "happy", "悲伤": "sad", "难过": "sad",
                   "激昂": "angry", "愤怒": "angry", "温柔": "gentle", "严肃": "serious"}.get(emo, "neutral")
        s["emotion"] = emo
        _sz = str(s.get("shot_size") or "").strip()
        _sz = _sz.replace("特写", "closeup").replace("近景", "medium").replace("中景", "medium").replace("全景", "wide").replace("远景", "wide")
        s["shot_size"] = _sz if _sz in ("closeup", "medium", "wide") else ""
        search = str(s.get("search") or "").strip()
        search = re.sub(r"[\"'\[\]]", "", search)[:60]
        if not search:
            search = (s.get("shot") or "").strip()[:30]
        s["search"] = search
    return {"title": data.get("title") or "未命名短剧", "episode": data.get("episode") or 1, "scenes": scenes, "characters": _parse_characters(data)}


@router.post("/novel-to-script")
async def novel_to_script(
    novel: str = Form(...),
    title: str = Form("", description="剧名（空=LLM 根据原文起名）"),
    duration: int = Form(120, description="单集目标时长（秒）"),
    episode: int = Form(1, description="第几集（>1 时参考上一集角色表保持一致性）"),
    series_id: str = Form("", description="所属系列 ID（连载剧集，复用角色库）"),
    current_user: dict = require_auth(),
):
    """小说原文 → 短剧剧本分镜（红果短剧风格）。

    - 粘贴小说/故事原文，LLM 提取核心角色表 + 分镜场景
    - series_id 指定系列时，自动合并该系列已保存的角色圣经（跨集角色一致）
    - 返回 scenes + characters，可直接进 /generate 或保存为系列
    """
    novel = novel.strip()
    if len(novel) < 30:
        raise HTTPException(400, "请粘贴更完整的小说/故事原文（至少 30 字）")
    duration_hint = max(20, min(1800, int(duration) or 120))
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""

    # 系列角色库：跨集一致性（角色圣经）
    series_chars = []
    if series_id:
        conn = get_db()
        try:
            _ensure_drama_tables(conn)
            rows = conn.execute(
                "SELECT * FROM drama_characters WHERE series_id=? AND user_id=? ORDER BY idx",
                (series_id, user),
            ).fetchall()
            series_chars = [dict(r) for r in rows]
        finally:
            conn.close()

    chars_block = ""
    if series_chars:
        chars_block = (
            "\n【本系列已确立的角色圣经（必须沿用，禁止改外貌服装）】\n"
            + "\n".join(
                f"- {c['name']}（{c['gender']} {c['age']}）：{c['appearance']}，{c['outfit']}；英文特征 {c['search']}"
                for c in series_chars
            )
        )

    ep_block = f"\n本集为第 {episode} 集，开头 5-10 秒要承接上一集结尾的悬念钩子。" if episode and int(episode) > 1 else ""

    # 长剧分块：>240s 时 LLM 单次输出 24+ 镜 JSON 极易格式错误，
    # 改为按「每块 ≤12 镜」分批生成再合并（每块独立 JSON，短输出更稳定）
    batch_count = 1
    if duration_hint > 240:
        batch_count = math.ceil(duration_hint / 240)  # 600s → 3 块，900s → 4 块
    scenes_all: list = []
    chars_all: list = []
    title_out = ""
    for bi in range(batch_count):
        batch_scenes = math.ceil((duration_hint / batch_count) / 28)  # 每块约 28s/镜
        start_idx = len(scenes_all) + 1
        block_prompt = (
            f"小说原文（可截取关键章节）：\n{novel[:12000]}\n\n"
            f"目标单集时长约 {duration_hint} 秒，共分 {batch_count} 块生成。"
            f"本块为第 {bi + 1}/{batch_count} 块（分镜序号从 {start_idx} 开始，"
            f"本块生成约 {batch_scenes} 镜）。"
            f"{chars_block}{ep_block}\n"
            f"第 1 块需输出完整 characters 角色表；后续块可省略 characters（沿用第 1 块）。"
            f"剧情需连贯：第 {start_idx} 镜承接上一块结尾，本块结尾留悬念。"
            f"\n重要：直接输出 JSON（不要 markdown 代码块、不要任何解释文字、不要省略号结尾）。"
        )
        ok_block = False
        for attempt in range(3):
            try:
                raw = await call_llm_async(
                    _NOVEL_SYSTEM,
                    block_prompt,
                    max_tokens=16000,
                    temperature=0.85,
                    timeout=300,
                )
                if not raw or len(raw.strip()) < 20:
                    raise ValueError("LLM 输出为空")
                partial = _drama_parse_script(raw)
                # 本块 scenes 的 id 从 1 起，需要偏移到全局序号
                for s in partial["scenes"]:
                    s["id"] = start_idx
                    start_idx += 1
                scenes_all.extend(partial["scenes"])
                if partial.get("characters"):
                    chars_all = partial["characters"]
                if not title_out:
                    title_out = partial["title"]
                ok_block = True
                break
            except (ValueError, json.JSONDecodeError) as e:
                logger.warning(f"剧本分块 {bi + 1} 解析失败: {e} | raw前200: {raw[:200]!r}")
                try:
                    with open(f"/tmp/raw_block_{bi}.txt", "w", encoding="utf-8") as _f:
                        _f.write(raw or "")
                except Exception:
                    pass
                if attempt == 2:
                    raise HTTPException(502, f"剧本解析失败（第 {bi + 1} 块）：{e}") from e
        if not ok_block:
            raise HTTPException(502, "剧本生成失败，请稍后重试")
    if not scenes_all:
        raise HTTPException(502, "剧本没有分镜")
    script = {"title": title_out or "未命名短剧", "episode": int(episode or 1), "scenes": scenes_all, "characters": chars_all}
    # 系列角色库优先：LLM 可能改了角色，用已存角色表覆盖（保证跨集稳定）
    if series_chars:
        script["characters"] = series_chars
    # 时长防御
    from short_drama import _enforce_duration as _drama_enforce

    script["scenes"] = _drama_enforce(script["scenes"], duration_hint)
    return {
        "title": title.strip() or script["title"],
        "episode": int(episode or 1),
        "scenes": script["scenes"],
        "characters": script.get("characters") or [],
        "series_id": series_id,
        "message": "小说已转为剧本分镜，可编辑后生成，或保存为系列连载",
    }


def _now() -> str:
    """当前时间 ISO。"""
    return datetime.now().isoformat()


def _ensure_drama_tables(conn) -> None:
    """红果短剧升级：系列 + 角色圣经表。"""
    conn.execute(
        """CREATE TABLE IF NOT EXISTS drama_series (
            id TEXT PRIMARY KEY,
            user_id TEXT DEFAULT '',
            name TEXT DEFAULT '',
            description TEXT DEFAULT '',
            genre TEXT DEFAULT '',          -- 题材：都市/甜宠/战神/逆袭…
            total_episodes INTEGER DEFAULT 0,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS drama_characters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id TEXT DEFAULT '',
            user_id TEXT DEFAULT '',
            idx INTEGER DEFAULT 0,
            cid TEXT DEFAULT '',           -- 角色 id（lin）
            name TEXT DEFAULT '',
            gender TEXT DEFAULT '',
            age TEXT DEFAULT '',
            appearance TEXT DEFAULT '',
            outfit TEXT DEFAULT '',
            search TEXT DEFAULT '',        -- 英文特征词（素材/图生图锚定）
            portrait_path TEXT DEFAULT '', -- 角色定妆参考图（本地路径）
            portrait_url TEXT DEFAULT '',  -- 参考图访问 URL
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )"""
    )


class SeriesCreateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=60, description="系列名")
    description: str = Field("", max_length=300)
    genre: str = Field("都市", max_length=20, description="题材")


class CharacterSaveRequest(BaseModel):
    characters: list[dict] = Field(default_factory=list, description="角色表（id/name/gender/age/appearance/outfit/search）")


@router.post("/series")
async def create_series(req: SeriesCreateRequest, current_user: dict = require_auth()):
    """创建短剧系列（连载容器）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_drama_tables(conn)
    try:
        sid = f"ds_{uuid.uuid4().hex[:10]}"
        conn.execute(
            "INSERT INTO drama_series (id, user_id, name, description, genre, created_at, updated_at) VALUES (?,?,?,?,?,?,?)",
            (sid, user, req.name.strip(), req.description.strip(), req.genre.strip(), _now(), _now()),
        )
        conn.commit()
        return {"ok": True, "id": sid, "message": f"系列「{req.name.strip()}」已创建"}
    finally:
        conn.close()


@router.get("/series")
async def list_series(current_user: dict = require_auth()):
    """短剧系列列表（含每系列角色数）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_drama_tables(conn)
    try:
        rows = conn.execute(
            "SELECT * FROM drama_series WHERE user_id=? ORDER BY created_at DESC", (user,)
        ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            cnt = conn.execute(
                "SELECT COUNT(*) c FROM drama_characters WHERE series_id=?", (d["id"],)
            ).fetchone()["c"]
            d["character_count"] = cnt
            out.append(d)
        return {"series": out}
    finally:
        conn.close()


@router.delete("/series/{series_id}")
async def delete_series(series_id: str, current_user: dict = require_auth()):
    """删除系列及其角色库。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_drama_tables(conn)
    try:
        row = conn.execute(
            "SELECT id FROM drama_series WHERE id=? AND user_id=?", (series_id, user)
        ).fetchone()
        if not row:
            raise HTTPException(404, "系列不存在")
        conn.execute("DELETE FROM drama_characters WHERE series_id=?", (series_id,))
        conn.execute("DELETE FROM drama_series WHERE id=?", (series_id,))
        conn.commit()
        return {"ok": True, "message": "系列已删除"}
    finally:
        conn.close()


@router.put("/series/{series_id}/characters")
async def save_series_characters(series_id: str, req: CharacterSaveRequest, current_user: dict = require_auth()):
    """保存系列角色圣经（跨集一致性核心）：覆盖该系列角色表。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_drama_tables(conn)
    try:
        row = conn.execute(
            "SELECT id FROM drama_series WHERE id=? AND user_id=?", (series_id, user)
        ).fetchone()
        if not row:
            raise HTTPException(404, "系列不存在")
        if len(req.characters) > 10:
            raise HTTPException(400, "单系列最多 10 个角色")
        conn.execute("DELETE FROM drama_characters WHERE series_id=?", (series_id,))
        now = _now()
        for i, c in enumerate(req.characters):
            if not isinstance(c, dict) or not c.get("name"):
                continue
            cid = re.sub(r"[^a-z0-9]", "", str(c.get("id") or "").strip().lower()) or f"c{i + 1}"
            conn.execute(
                """INSERT INTO drama_characters (series_id, user_id, idx, cid, name, gender, age, appearance, outfit, search, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    series_id, user, i, cid,
                    str(c.get("name") or "")[:30], str(c.get("gender") or "女")[:10],
                    str(c.get("age") or "")[:10], str(c.get("appearance") or "")[:100],
                    str(c.get("outfit") or "")[:100], str(c.get("search") or "")[:80],
                    now, now,
                ),
            )
        conn.commit()
        return {"ok": True, "saved": len(req.characters), "message": f"已保存 {len(req.characters)} 个角色圣经"}
    finally:
        conn.close()


@router.get("/series/{series_id}/characters")
async def get_series_characters(series_id: str, current_user: dict = require_auth()):
    """读取系列角色圣经（跨集复用）。"""
    user = current_user.get("username", "") if isinstance(current_user, dict) else ""
    conn = get_db()
    _ensure_drama_tables(conn)
    try:
        rows = conn.execute(
            "SELECT * FROM drama_characters WHERE series_id=? AND user_id=? ORDER BY idx",
            (series_id, user),
        ).fetchall()
        return {"characters": [dict(r) for r in rows]}
    finally:
        conn.close()
